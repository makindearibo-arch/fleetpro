#!/usr/bin/env python
r"""
FleetPro: Import AKURE 4 daily diesel readings (Jan 1 - Jul 7 2026).

Source: "AKURE 4 Daily Diesel TEMPLATE YEAR 2026.xlsx" - monthly tabs
JANUARY 2026..JULY 2026 among many unrelated trackers. Narrow layout
(same as Ondo 2 / Ado 1): date col 1, supplier 2, opening 4, purchases 5,
transfer 6, closing 8/9, consumption 10, gen meter 12/13, NEPA 16/17.

METER IS IN KILO-HOURS: the DCP-10 panel reads e.g. "15.62 Kh" and the
sheet records it verbatim (14.67 Jan 1 -> 15.62 Jul 7 = ~950 real hours).
Values are imported AS-SHEET (Kh) so the stored meter matches what staff
see and type on the panel - converting to hours would make the app's
locked opening (15,610) reject the staff's typed closing (15.62) via the
backwards-meter guard. Consequences: hours_run is ~0.00-0.02/day (0.01 Kh
= 10 h resolution), so per-hour rates are meaningless - consumption_rate
is stored NULL and no baseline will be learned (min-burn floor keeps the
store from flagging). Akure 4's tracking rides on its LEVEL data, which
is clean daily (10-200 L/day; balance verified).

The store manager's 7 app entries (Jun 30 - Jul 6) typed TANK LEVELS into
the hours fields (550->550, 540->540) yielding negative consumption; the
sheet covers those dates, so --replace deletes them (per the user).

Usage:
  py scripts\import_akure4.py                       # dry run
  py scripts\import_akure4.py --apply --replace     # sheet is authoritative

Then: sync_generator_hours.py --apply (gen.hrs -> 15.62, matching the panel).
"""
import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
import datetime
from pathlib import Path

import openpyxl

FILE = Path(r"C:\Users\MakindeAribo\Downloads\AKURE 4 Daily Diesel TEMPLATE YEAR 2026.xlsx")
STORE_NAME = "Akure 4"
GEN_LOC_MATCH = "akure 4"
YEAR = 2026

MONTH_TABS = {"JANUARY 2026": 1, "FEBRUARY 2026": 2, "MARCH 2026": 3, "APRIL 2026": 4,
              "MAY 2026": 5, "JUNE 2026": 6, "JULY 2026": 7}

OFF = {
    "supplier": 2, "opening_total": 4, "purchases_in": 5, "transfer_out": 6,
    "closing_main": 8, "closing_total": 9, "consumption_l": 10,
    "gen_h_open": 12, "gen_h_close": 13, "hour_run": 14, "consumption_rate": 15,
    "nepa_open": 16, "nepa_close": 17,
}


def load_env_file():
    for name in (".env", "SupabaseCreds.env", "supabase.env"):
        p = Path(name)
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
        return


class Supabase:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def _req(self, method, path, params=None, body=None, extra_headers=None):
        url = f"{self.url}/rest/v1/{path}"
        if params:
            url += "?" + urllib.parse.urlencode(params, doseq=True)
        data = json.dumps(body).encode("utf-8") if body is not None else None
        headers = dict(self.headers)
        if extra_headers:
            headers.update(extra_headers)
        req = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read()
                return json.loads(raw) if raw else []
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {e.read().decode('utf-8','replace')}") from e

    def select(self, table, columns="*", filters=None):
        params = {"select": columns}
        if filters:
            params.update(filters)
        return self._req("GET", table, params=params)

    def insert(self, table, rows):
        return self._req("POST", table, body=rows, extra_headers={"Prefer": "return=representation"})

    def delete(self, table, filters):
        return self._req("DELETE", table, params=filters, extra_headers={"Prefer": "return=minimal"})


def n(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return None if (isinstance(v, float) and v != v) else float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("#"):
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


def detect_date_col(ws):
    for dc in (1, 2):
        for r in range(1, ws.max_row + 1):
            if isinstance(ws.cell(r, dc).value, datetime.datetime):
                return dc
    return None


def parse_readings():
    wb = openpyxl.load_workbook(FILE, data_only=True)
    out = []
    for tab, month in MONTH_TABS.items():
        if tab not in wb.sheetnames:
            print(f"  ! missing tab {tab!r}")
            continue
        ws = wb[tab]
        dc = detect_date_col(ws)
        if dc is None:
            continue
        base = dc - 1
        def col(key):
            return base + OFF[key]
        cnt = 0
        for r in range(1, ws.max_row + 1):
            d = ws.cell(r, dc).value
            if not isinstance(d, datetime.datetime) or d.year != YEAR or d.month != month:
                continue
            closing = n(ws.cell(r, col("closing_total")).value)
            if closing is None or closing <= 0:
                closing = n(ws.cell(r, col("closing_main")).value)
            gh_close = n(ws.cell(r, col("gen_h_close")).value)
            if not ((closing is not None and closing > 0) or (gh_close is not None and gh_close > 0)):
                continue
            gh_open = n(ws.cell(r, col("gen_h_open")).value)
            cons = n(ws.cell(r, col("consumption_l")).value)
            hours = (gh_close - gh_open) if (gh_open is not None and gh_close is not None) else None
            rate = None  # meter is in Kh (10 h resolution) - per-hour rates are meaningless
            sup = ws.cell(r, col("supplier")).value
            sup = sup.strip() if isinstance(sup, str) and sup.strip() else None
            out.append({
                "date": d.date().isoformat(),
                "gen_h_open": gh_open, "gen_h_close": gh_close,
                "closing": round(closing, 1) if closing is not None else None,
                "purchases_in": n(ws.cell(r, col("purchases_in")).value),
                "consumption_l": round(cons, 1) if cons is not None else None,
                "consumption_rate": rate,
                "nepa_open": n(ws.cell(r, col("nepa_open")).value),
                "nepa_close": n(ws.cell(r, col("nepa_close")).value),
                "supplier": sup,
            })
            cnt += 1
        print(f"  {tab:<10} datecol={dc}: {cnt} rows")
    out.sort(key=lambda r: r["date"])
    return out


def recalc_baseline(sb, gen_id):
    rows = sb._req("GET", "diesel_readings", params={
        "select": "date,gen_hours_opening,gen_hours_closing,diesel_level_actual,diesel_added",
        "generator_id": f"eq.{gen_id}", "order": "date.asc",
    })
    rates = []
    prev = None
    prev_date = None
    for r in rows:
        ho, hc, lvl = r.get("gen_hours_opening"), r.get("gen_hours_closing"), r.get("diesel_level_actual")
        added = r.get("diesel_added") or 0
        hrs = (hc - ho) if (ho is not None and hc is not None) else None
        cur = datetime.date.fromisoformat(r["date"]) if r.get("date") else None
        consec = prev_date is not None and cur is not None and (cur - prev_date).days == 1
        if hrs and hrs > 0 and lvl is not None and prev is not None and consec:
            actual = prev + added - lvl
            if actual > 0:
                rate = actual / hrs
                if 1 <= rate <= 100:
                    rates.append(rate)
        if lvl is not None:
            prev = lvl
            prev_date = cur
    if not rates:
        print("  No valid rate pairs — skipping baseline.")
        return
    avg = sum(rates) / len(rates)
    flagged = sum(1 for r in rates if abs(r - avg) / avg > 0.20) / len(rates) * 100
    print(f"  Pairs: {len(rates)}  Avg: {avg:.2f} L/hr (min {min(rates):.2f}, max {max(rates):.2f})  ~{flagged:.1f}% >20% off mean")


def main(apply_mode, recalc_mode, replace_mode):
    load_env_file()
    url, key = os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY required.")
        sys.exit(1)
    sb = Supabase(url, key)

    print("=== PARSING AKURE 4 ===")
    readings = parse_readings()
    print(f"  Total: {len(readings)} reading rows\n")

    print("=== SUPABASE LOOKUP ===")
    gens = sb.select("generators", columns="id,name,loc,asset_type")
    gen = next((g for g in gens if (g.get("loc") or "").strip().lower() == GEN_LOC_MATCH
                and (g.get("asset_type") or "generator") != "oven"), None)
    if not gen:
        print(f"ERROR: no generator with loc == {STORE_NAME!r}. Found locs:")
        for g in sorted(gens, key=lambda x: (x.get("loc") or "")):
            print(f"  - {g['name']!r} loc={g.get('loc')!r}")
        sys.exit(1)
    gen_id = gen["id"]
    print(f"  Akure 4 generator: {gen['name']} (id={gen_id}, loc={gen.get('loc')})")
    existing = sb.select("diesel_readings", columns="id,date,submitted_by", filters={"generator_id": f"eq.{gen_id}"})
    existing_dates = {r["date"] for r in existing}
    print(f"  Existing readings for this generator: {len(existing_dates)}")

    sheet_dates = {r["date"] for r in readings}
    overlap = [r for r in existing if r["date"] in sheet_dates]

    if replace_mode:
        new_readings = readings
        print(f"\n=== REPLACE MODE ===")
        print(f"  Will DELETE {len(overlap)} existing reading(s) on sheet dates, then insert all {len(readings)}.")
        for r in sorted(overlap, key=lambda x: x["date"]):
            print(f"    - delete {r['date']} (app entry by {r.get('submitted_by')})")
    else:
        new_readings = [r for r in readings if r["date"] not in existing_dates]

    print(f"\n=== INSERT PLAN ===")
    print(f"  diesel_readings to insert: {len(new_readings)} "
          f"({'replace mode — all sheet rows' if replace_mode else f'skip {len(readings) - len(new_readings)} dupes'})")
    if new_readings:
        print(f"  Date range: {new_readings[0]['date']} .. {new_readings[-1]['date']}")
        print(f"  Diesel added (purchases/transfer-in) total: {sum(r['purchases_in'] or 0 for r in new_readings):,.0f} L")

    if not apply_mode:
        print("\nDRY RUN — no writes. Re-run with --apply (add --recalc-baseline for a baseline preview).")
        if recalc_mode:
            print("\n=== BASELINE PREVIEW (gap-aware) ===")
            recalc_baseline(sb, gen_id)
        return

    print("\n=== WRITING ===")
    if replace_mode and overlap:
        for r in overlap:
            sb.delete("diesel_readings", {"id": f"eq.{r['id']}"})
        print(f"  - Deleted {len(overlap)} overlapping reading(s).")
    if new_readings:
        payload = [{
            "generator_id": gen_id, "store_location": STORE_NAME, "date": r["date"],
            "gen_hours_opening": r["gen_h_open"], "gen_hours_closing": r["gen_h_close"],
            "diesel_level_actual": r["closing"], "diesel_added": r["purchases_in"] or 0,
            "consumption_litres": r["consumption_l"],
            "consumption_rate": r["consumption_rate"],
            "nepa_meter_opening": r["nepa_open"], "nepa_meter_closing": r["nepa_close"],
            "submitted_by": None,
            "notes": f"Supplier: {r['supplier']}" if r["supplier"] else None,
        } for r in new_readings]
        BATCH = 100
        total = 0
        for i in range(0, len(payload), BATCH):
            total += len(sb.insert("diesel_readings", payload[i:i + BATCH]))
        print(f"  + Inserted {total} Akure 4 readings")
    print("\nDONE. Next: recalc_baselines.py --apply -> backfill_discrepancy_flags.py --apply -> sync_generator_hours.py --apply")
    if recalc_mode:
        print("\n=== BASELINE PREVIEW (gap-aware) ===")
        recalc_baseline(sb, gen_id)


if __name__ == "__main__":
    main("--apply" in sys.argv, "--recalc-baseline" in sys.argv, "--replace" in sys.argv)
