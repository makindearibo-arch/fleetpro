#!/usr/bin/env python
r"""
FleetPro: Import Okitipupa CR daily diesel readings (DIESEL TRACKER 2026.xlsx).

Okitipupa CR has TWO generators. The hour meter resets from ~25,277 (end April)
to ~145 (start May) because a second/newer generator took over:

  Jan-April readings  -> 'Okitipupa 2nd Generator' (G-010)   [old unit, ~25k hrs]
  May + June readings -> 'Okitipupa CR Generator'            [new unit, ~145 hrs]

The 'JUNE' tab rows were mis-dated by the employee as May 1-2; they are actually
June 1-2 and are re-dated here.

Column layout is identical to OWO CR (no 'RECEIVED BY' column).

Usage:
  py scripts\import_okitipupa_cr.py                       # dry run
  py scripts\import_okitipupa_cr.py --apply               # write
  py scripts\import_okitipupa_cr.py --apply --recalc-baseline

Reads SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY from .env / SupabaseCreds.env / env.
"""
import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
import datetime
from pathlib import Path

import openpyxl

FILE = Path(r"C:\Users\MakindeAribo\Downloads\DIESEL TRACKER 2026 (1).xlsx")
STORE_NAME = "Okitipupa CR"

# Generator records (both must exist with loc 'Okitipupa CR')
GEN_A_NAME = "Okitipupa 2nd Generator"   # Jan-April (old unit)
GEN_B_NAME = "Okitipupa CR Generator"    # May + June (new unit)

# tab -> (year, month, gen_key, force_month)
#   force_month=True overrides the cell's month with the config month (JUNE tab fix)
TABS = {
    "JANUARY":  (2026, 1, "A", False),
    "FEBRUARY": (2026, 2, "A", False),
    "MARCH":    (2026, 3, "A", False),
    "APRIL":    (2026, 4, "A", False),
    "MAY":      (2026, 5, "B", False),
    "JUNE":     (2026, 6, "B", True),   # rows mis-dated May 1-2 -> June 1-2
}

# OWO CR / Okitipupa CR column map (1-indexed)
COL = {
    "date": 1, "supplier": 2, "opening_main": 3, "purchases_in": 5, "transfer_out": 6,
    "closing_main": 8, "consumption_l": 10, "gen_h_open": 12, "gen_h_close": 13,
    "consumption_rate": 15, "nepa_open": 16, "nepa_close": 17,
}


def load_env_file():
    for name in (".env", "SupabaseCreds.env", "supabase.env"):
        p = Path(name)
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
        return


class Supabase:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def _req(self, method, path, params=None, body=None, extra_headers=None):
        full = f"{self.url}/rest/v1/{path}"
        if params:
            full += "?" + urllib.parse.urlencode(params, doseq=True)
        data = json.dumps(body).encode("utf-8") if body is not None else None
        headers = dict(self.headers)
        if extra_headers:
            headers.update(extra_headers)
        req = urllib.request.Request(full, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read()
                return json.loads(raw) if raw else []
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {e.read().decode('utf-8','replace')}") from e

    def select(self, table, columns="*", filters=None):
        params = {"select": columns}
        if filters:
            params.update(filters)
        return self._req("GET", table, params=params)

    def insert(self, table, rows):
        return self._req("POST", table, body=rows, extra_headers={"Prefer": "return=representation"})


def n(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return None if (isinstance(v, float) and v != v) else float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("#"):
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


def parse():
    wb = openpyxl.load_workbook(FILE, data_only=True)
    out = {"A": [], "B": []}
    for tab, (year, month, gen_key, force) in TABS.items():
        if tab not in wb.sheetnames:
            print(f"  ! missing tab {tab!r}")
            continue
        ws = wb[tab]
        start = None
        for r in range(1, ws.max_row + 1):
            if isinstance(ws.cell(r, 1).value, datetime.datetime):
                start = r
                break
        if start is None:
            print(f"  ! no data rows in {tab}")
            continue
        cnt = 0
        for r in range(start, ws.max_row + 1):
            d = ws.cell(r, COL["date"]).value
            if not isinstance(d, datetime.datetime):
                continue
            gho = n(ws.cell(r, COL["gen_h_open"]).value)
            ghc = n(ws.cell(r, COL["gen_h_close"]).value)
            cs = n(ws.cell(r, COL["closing_main"]).value)
            # Skip empty template rows BEFORE date handling (blank month-end rows
            # would otherwise break the force re-date). A real completed day must
            # have a closing tank level OR closing gen hours (> 0).
            if not ((cs is not None and cs > 0) or (ghc is not None and ghc > 0)):
                continue
            if force:
                try:
                    use_date = datetime.date(year, month, d.day)
                except ValueError:
                    print(f"  ! {tab} row {r}: day {d.day} invalid for {year}-{month:02d}, skipping")
                    continue
            else:
                if d.year != year or d.month != month:
                    continue
                use_date = d.date()
            rec = {
                "date": use_date.isoformat(),
                "gen_h_open": gho,
                "gen_h_close": ghc,
                "closing_main": cs,
                "purchases_in": n(ws.cell(r, COL["purchases_in"]).value),
                "consumption_l": n(ws.cell(r, COL["consumption_l"]).value),
                "consumption_rate": n(ws.cell(r, COL["consumption_rate"]).value),
                "nepa_open": n(ws.cell(r, COL["nepa_open"]).value),
                "nepa_close": n(ws.cell(r, COL["nepa_close"]).value),
                "supplier": (ws.cell(r, COL["supplier"]).value or "").strip() if isinstance(ws.cell(r, COL["supplier"]).value, str) else None,
            }
            out[gen_key].append(rec)
            cnt += 1
        print(f"  {tab}: {cnt} rows -> Gen {gen_key}")
    return out


def recalc_baseline(sb, gen_id, label):
    rows = sb._req("GET", "diesel_readings", params={
        "select": "date,gen_hours_opening,gen_hours_closing,diesel_level_actual,diesel_added",
        "generator_id": f"eq.{gen_id}", "order": "date.asc"})
    rates = []
    prev = None
    for r in rows:
        ho, hc, lvl = r.get("gen_hours_opening"), r.get("gen_hours_closing"), r.get("diesel_level_actual")
        added = r.get("diesel_added") or 0
        hrs = (hc - ho) if (ho is not None and hc is not None) else None
        if hrs and hrs > 0 and lvl is not None and prev is not None:
            actual = prev + added - lvl
            if actual > 0:
                rate = actual / hrs
                if 1 <= rate <= 100:
                    rates.append(rate)
        if lvl is not None:
            prev = lvl
    if not rates:
        print(f"  [{label}] no valid rate pairs — skip baseline")
        return
    avg = sum(rates) / len(rates)
    flagged = sum(1 for x in rates if abs(x - avg) / avg > 0.20) / len(rates) * 100
    print(f"  [{label}] pairs={len(rates)} avg={avg:.2f} L/hr (min {min(rates):.2f}, max {max(rates):.2f}) ~{flagged:.1f}% >20% off")
    sb._req("POST", "generator_baselines", params={"on_conflict": "generator_id"},
            body=[{"generator_id": gen_id, "avg_litres_per_hour": round(avg, 2),
                   "baseline_readings_count": len(rates),
                   "last_calculated": datetime.datetime.now(datetime.timezone.utc).isoformat(),
                   "min_rate": round(min(rates), 2), "max_rate": round(max(rates), 2)}],
            extra_headers={"Prefer": "resolution=merge-duplicates,return=representation"})
    print(f"  [{label}] baseline upserted")


def build_payload(recs, gen_id):
    out = []
    for r in recs:
        out.append({
            "generator_id": gen_id, "store_location": STORE_NAME, "date": r["date"],
            "gen_hours_opening": r["gen_h_open"], "gen_hours_closing": r["gen_h_close"],
            "diesel_level_actual": r["closing_main"], "diesel_added": r["purchases_in"] or 0,
            "consumption_litres": int(r["consumption_l"]) if r["consumption_l"] is not None else None,
            "consumption_rate": r["consumption_rate"],
            "nepa_meter_opening": r["nepa_open"], "nepa_meter_closing": r["nepa_close"],
            "submitted_by": None,
            "notes": f"Supplier: {r['supplier']}" if r["supplier"] else None,
        })
    return out


def main(apply_mode, recalc_mode):
    load_env_file()
    url, key = os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY required.")
        sys.exit(1)
    sb = Supabase(url, key)

    print("=== PARSING Okitipupa CR ===")
    parsed = parse()
    print(f"  Gen A (Jan-Apr): {len(parsed['A'])} rows | Gen B (May-Jun): {len(parsed['B'])} rows\n")

    print("=== SUPABASE LOOKUP ===")
    gens = sb.select("generators", columns="id,name,loc")
    okit = [g for g in gens if (g.get("loc") or "").strip().lower() == "okitipupa cr"]
    gen_a = next((g for g in okit if (g["name"] or "").strip().lower() == GEN_A_NAME.lower()), None)
    gen_b = next((g for g in okit if (g["name"] or "").strip().lower() == GEN_B_NAME.lower()), None)
    if not gen_a or not gen_b:
        print(f"ERROR: need both '{GEN_A_NAME}' and '{GEN_B_NAME}' at loc 'Okitipupa CR'.")
        print("  Okitipupa CR generators found:")
        for g in okit:
            print(f"    - {g['name']!r} id={g['id']}")
        print("  (Run setup_store_generators.py --apply first if the 2nd one is missing.)")
        sys.exit(1)
    print(f"  Gen A: {gen_a['name']} (id={gen_a['id']})")
    print(f"  Gen B: {gen_b['name']} (id={gen_b['id']})")

    def existing_dates(gid):
        rows = sb.select("diesel_readings", columns="date", filters={"generator_id": f"eq.{gid}"})
        return {r["date"] for r in rows}

    ex_a, ex_b = existing_dates(gen_a["id"]), existing_dates(gen_b["id"])
    new_a = [r for r in parsed["A"] if r["date"] not in ex_a]
    new_b = [r for r in parsed["B"] if r["date"] not in ex_b]

    print("\n=== INSERT PLAN ===")
    print(f"  Gen A ({gen_a['name']}): {len(new_a)} new (skip {len(parsed['A'])-len(new_a)} dupes)")
    if new_a:
        print(f"     dates {new_a[0]['date']} .. {new_a[-1]['date']}")
    print(f"  Gen B ({gen_b['name']}): {len(new_b)} new (skip {len(parsed['B'])-len(new_b)} dupes)")
    if new_b:
        print(f"     dates {new_b[0]['date']} .. {new_b[-1]['date']}")

    if not apply_mode:
        print("\nDRY RUN — no writes. Re-run with --apply to write.")
        if recalc_mode:
            print("\n=== BASELINE RECALC ===")
            recalc_baseline(sb, gen_a["id"], gen_a["name"])
            recalc_baseline(sb, gen_b["id"], gen_b["name"])
        return

    print("\n=== WRITING ===")
    for label, gid, recs in [("A", gen_a["id"], new_a), ("B", gen_b["id"], new_b)]:
        payload = build_payload(recs, gid)
        BATCH = 100
        total = 0
        for i in range(0, len(payload), BATCH):
            total += len(sb.insert("diesel_readings", payload[i:i + BATCH]))
        print(f"  + Gen {label}: inserted {total} readings")
    print("\nDONE.")
    if recalc_mode:
        print("\n=== BASELINE RECALC ===")
        recalc_baseline(sb, gen_a["id"], gen_a["name"])
        recalc_baseline(sb, gen_b["id"], gen_b["name"])


if __name__ == "__main__":
    main("--apply" in sys.argv, "--recalc-baseline" in sys.argv)
