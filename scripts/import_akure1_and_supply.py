#!/usr/bin/env python
"""
FleetPro: Import Akure 1 daily diesel readings + admin diesel purchases/distributions.

Talks to Supabase via its REST API (PostgREST) using only stdlib + openpyxl.

Usage:
  # Dry-run (prints what would be inserted, doesn't write):
  py scripts\import_akure1_and_supply.py

  # Apply (actually writes to Supabase):
  py scripts\import_akure1_and_supply.py --apply

Env vars required (place in .env or set in shell):
  SUPABASE_URL                = https://bddmsrbfygbuyfdpieyl.supabase.co
  SUPABASE_SERVICE_ROLE_KEY   = <service_role key from Supabase Dashboard → Project Settings → API>

Install deps:
  py -m pip install openpyxl
  (No supabase / no python-dotenv needed — we parse .env manually below.)
"""
import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
import datetime
from pathlib import Path

import openpyxl

# ---------- Config ----------
AKURE_FILE = Path(r"C:\Users\MakindeAribo\Downloads\AKURE 1 DIESEL REPORT.xlsx")
SUPPLY_FILE = Path(r"C:\Users\MakindeAribo\Downloads\DIESEL REPORT TEMPLATE (1).xlsx")

AKURE_STORE_NAME = "Akure 1"
AKURE_GEN_NAME_HINT = "akure 1"

AKURE_MONTHLY_TABS = [
    "MARCH 2025", "APRIL 2025", "MAY 2025", "JUNE 2025", "JULY 2025", "AUGUST 2025",
    "SEPTEMBER 2025", "OCTOBER 2025", "NOVEMBER 2025", "DECEMBER 2025",
    "JANUARY 2026", "FEBRUARY 2026", "MARCH 2026", "APRIL", "MAY",
]
TAB_TO_YEAR_MONTH = {
    "MARCH 2025": (2025, 3), "APRIL 2025": (2025, 4), "MAY 2025": (2025, 5),
    "JUNE 2025": (2025, 6), "JULY 2025": (2025, 7), "AUGUST 2025": (2025, 8),
    "SEPTEMBER 2025": (2025, 9), "OCTOBER 2025": (2025, 10), "NOVEMBER 2025": (2025, 11),
    "DECEMBER 2025": (2025, 12),
    "JANUARY 2026": (2026, 1), "FEBRUARY 2026": (2026, 2), "MARCH 2026": (2026, 3),
    "APRIL": (2026, 4), "MAY": (2026, 5),
}

STORE_NORMALIZE = {
    "AKUNGBA": "Akungba CR", "AKUNGBA C.R.": "Akungba CR",
    "AKURE 1": "Akure 1", "AKURE 2": "Akure 2", "AKURE 4": "Akure 4",
    "AKURE 5": "Akure 5", "AKURE 6": "Akure 6", "AKURE BAKERY": "Akure Bakery",
    "ALAGBAKA GEN": "Akure 1", "ALAGBAKA GEN(OVERAGE)": "Akure 1",
    "OVERAGE - AKURE 1": "Akure 1",
    "IDANRE": "Idanre CR", "IGBOKODA": "Igbokoda CR",
    "IKARE BAKERY": "Ikare Bakery", "IKARE BAKEKRY": "Ikare Bakery", "IKARE BAKKERY": "Ikare Bakery",
    "IKARE C.R.": "Ikare CR", "IKARE CR": "Ikare CR",
    "OKITIPUPA BAKERY": "Okitipupa Bakery",
    "OKITIPUPA C.R": "Okitipupa CR", "OKITIPUPA C.R.": "Okitipupa CR", "OKITIPUPA CR": "Okitipupa CR",
    "OND0 1": "Ondo CR", "ONDO 1": "Ondo CR",
    "ONDO 2": "Ondo 2 CR", "ONDO CR 2": "Ondo 2 CR",
    "ONDO BAKERY": "Ondo Bakery",
    "OWO BAKERY": "Owo Bakery", "OWO C.R.": "Owo CR", "OWO CR": "Owo CR",
    "PIE EXPRESS/WAREHOUSE": "Pie Express / Warehouse",
}
OVERAGE_SOURCES = {"ALAGBAKA GEN(OVERAGE)", "OVERAGE - AKURE 1"}
SKIP_STORES = {"BALANCE", "STORE SUPPLIED", None, ""}


# ---------- .env loader (no python-dotenv dep) ----------
def load_env_file():
    # Try common filenames in order; first match wins.
    for name in (".env", "SupabaseCreds.env", "supabase.env"):
        p = Path(name)
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            v = v.strip().strip('"').strip("'")
            os.environ.setdefault(k.strip(), v)
        return


# ---------- Supabase REST client (stdlib only) ----------
class Supabase:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.key = key
        self.headers = {
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
        }

    def _req(self, method, path, params=None, body=None, extra_headers=None):
        url = f"{self.url}/rest/v1/{path}"
        if params:
            url += "?" + urllib.parse.urlencode(params, doseq=True)
        data = None
        if body is not None:
            data = json.dumps(body).encode("utf-8")
        headers = dict(self.headers)
        if extra_headers:
            headers.update(extra_headers)
        req = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read()
                if not raw:
                    return []
                return json.loads(raw)
        except urllib.error.HTTPError as e:
            msg = e.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {msg}") from e

    def select(self, table, columns="*", filters=None, limit=None):
        params = {"select": columns}
        if filters:
            params.update(filters)
        if limit:
            params["limit"] = str(limit)
        return self._req("GET", table, params=params)

    def insert(self, table, rows):
        if not rows:
            return []
        return self._req("POST", table, body=rows,
                         extra_headers={"Prefer": "return=representation"})


# ---------- Helpers ----------
def fix_purchase_date(v):
    if isinstance(v, datetime.datetime):
        if v.day <= 12 and v.month <= 12 and v.day != v.month:
            try:
                return v.replace(month=v.day, day=v.month)
            except ValueError:
                return v
        return v
    if isinstance(v, str):
        s = v.strip().replace("-", "/")
        for fmt in ("%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None


def detect_data_start_row(ws):
    for r in range(1, ws.max_row + 1):
        if isinstance(ws.cell(r, 1).value, datetime.datetime):
            return r
    return None


def n(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (v != v):
            return None
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("#"):
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


# ---------- Parsers ----------
def parse_akure1_readings():
    wb = openpyxl.load_workbook(AKURE_FILE, data_only=True)
    out = []
    for tab in AKURE_MONTHLY_TABS:
        if tab not in wb.sheetnames:
            print(f"  ! sheet missing: {tab!r}")
            continue
        ws = wb[tab]
        start = detect_data_start_row(ws)
        if start is None:
            print(f"  ! no data rows found in {tab}")
            continue
        year, month = TAB_TO_YEAR_MONTH[tab]
        rows_added = 0
        for r in range(start, ws.max_row + 1):
            d = ws.cell(r, 1).value
            if not isinstance(d, datetime.datetime):
                continue
            if d.year != year or d.month != month:
                continue
            rec = {
                "date": d.date().isoformat(),
                "supplier": (ws.cell(r, 2).value or "").strip() if isinstance(ws.cell(r, 2).value, str) else None,
                "purchases_in": n(ws.cell(r, 5).value),
                "received_by": ws.cell(r, 7).value,
                "closing_main": n(ws.cell(r, 9).value),
                "consumption_l": n(ws.cell(r, 11).value),
                "gen_h_open": n(ws.cell(r, 13).value),
                "gen_h_close": n(ws.cell(r, 14).value),
                "consumption_rate": n(ws.cell(r, 16).value),
                "nepa_open": n(ws.cell(r, 17).value),
                "nepa_close": n(ws.cell(r, 18).value),
            }
            if rec["gen_h_open"] is None and rec["gen_h_close"] is None and rec["closing_main"] is None:
                continue
            out.append(rec)
            rows_added += 1
        print(f"  {tab}: {rows_added} rows")
    return out


def parse_purchases():
    wb = openpyxl.load_workbook(SUPPLY_FILE, data_only=True)
    ws = wb["DIESEL PURCHASE"]
    out = []
    for r in range(2, ws.max_row + 1):
        sn_val = ws.cell(r, 1).value
        if sn_val == "TOTAL" or not isinstance(sn_val, int):
            continue
        date_raw = ws.cell(r, 3).value
        litres = n(ws.cell(r, 4).value)
        supplier = ws.cell(r, 5).value
        price_per_l = n(ws.cell(r, 6).value)
        total = n(ws.cell(r, 7).value)
        if date_raw is None:
            continue
        d = fix_purchase_date(date_raw)
        if d is None:
            print(f"  ! could not parse purchase date at row {r}: {date_raw!r}")
            continue
        out.append({
            "date": d.date().isoformat(),
            "supplier": (supplier or "").strip() if isinstance(supplier, str) else "",
            "litres": litres or 0,
            "price_per_litre": price_per_l,
            "total_cost": total,
            "_sn": sn_val,
        })
    return out


def parse_distributions(purchases):
    wb = openpyxl.load_workbook(SUPPLY_FILE, data_only=True)
    ws = wb["STORES SUPPLY"]
    out = []
    last_date = None
    last_supplier = None
    for r in range(3, ws.max_row + 1):
        d_raw = ws.cell(r, 1).value
        if isinstance(d_raw, datetime.datetime):
            last_date = d_raw.date()
        supplier = ws.cell(r, 4).value
        qty_raw = ws.cell(r, 5).value
        store_raw = ws.cell(r, 6).value
        if supplier:
            last_supplier = str(supplier).strip()
        qty = n(qty_raw)
        if not qty or qty <= 0:
            continue
        store_key = (str(store_raw).strip() if store_raw is not None else "")
        if store_key in SKIP_STORES or store_key.upper() in {"BALANCE", "STORE SUPPLIED"}:
            continue
        mapped_store = STORE_NORMALIZE.get(store_key.upper().strip())
        if not mapped_store:
            print(f"  ! unmapped store at row {r}: {store_key!r}")
            continue
        is_overage = store_key.upper().strip() in OVERAGE_SOURCES
        if last_date is None:
            continue
        matched_purchase_sn = None
        if last_supplier:
            candidates = [p for p in purchases
                          if p["supplier"].upper() == last_supplier.upper()
                          and p["date"] <= last_date.isoformat()]
            if candidates:
                matched_purchase_sn = max(candidates, key=lambda p: p["date"])["_sn"]
        out.append({
            "date": last_date.isoformat(),
            "store_location": mapped_store,
            "litres": qty,
            "supplier": last_supplier,
            "purchase_sn": matched_purchase_sn,
            "is_overage": is_overage,
            "raw_store": store_key,
        })
    return out


# ---------- Main ----------
def main(apply_mode):
    load_env_file()
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY must be set in .env or env vars.")
        sys.exit(1)
    sb = Supabase(url, key)

    print("\n=== PARSING ===")
    print("Akure 1 monthly readings:")
    readings = parse_akure1_readings()
    print(f"  Total: {len(readings)} reading rows\n")

    print("Diesel purchases:")
    purchases = parse_purchases()
    for p in purchases:
        print(f"  {p['date']}  {p['litres']:>6.0f}L  {p['supplier']:<15}  @{p['price_per_litre'] or 0:>5}")
    print(f"  Total: {len(purchases)} purchases\n")

    print("Distributions:")
    distributions = parse_distributions(purchases)
    by_store = {}; overage_count = 0
    for d in distributions:
        by_store[d["store_location"]] = by_store.get(d["store_location"], 0) + 1
        if d["is_overage"]: overage_count += 1
    for store, c in sorted(by_store.items()):
        print(f"  {store:<30} {c} distributions")
    print(f"  Total: {len(distributions)} distributions ({overage_count} marked overage)\n")

    print("=== SUPABASE LOOKUP ===")
    gens = sb.select("generators", columns="id,name,loc")
    exact_loc = [g for g in gens if (g.get("loc") or "").strip().lower() == "akure 1"]
    gen = exact_loc[0] if exact_loc else next(
        (g for g in gens if AKURE_GEN_NAME_HINT in (g.get("name") or "").lower()), None)
    if not gen:
        print("ERROR: could not find Akure 1 generator. Add one in FleetPro before importing.")
        print("Generators in DB:")
        for g in gens:
            print(f"  - {g['name']!r} loc={g.get('loc')!r}")
        sys.exit(1)
    print(f"  Akure 1 generator: {gen['name']} (id={gen['id']}, loc={gen.get('loc')})")
    gen_id = gen["id"]

    existing_readings = sb.select("diesel_readings", columns="date",
                                  filters={"generator_id": f"eq.{gen_id}"})
    existing_dates = {r["date"] for r in existing_readings}
    print(f"  Existing readings for this generator: {len(existing_dates)}")

    existing_p = sb.select("diesel_purchases", columns="date,supplier,litres")
    existing_pkeys = {(r["date"], (r["supplier"] or "").upper(), float(r["litres"])) for r in existing_p}
    print(f"  Existing purchases in DB: {len(existing_pkeys)}")

    existing_d = sb.select("diesel_distributions", columns="date,store_location,litres")
    existing_dkeys = {(r["date"], r["store_location"], float(r["litres"])) for r in existing_d}
    print(f"  Existing distributions in DB: {len(existing_dkeys)}")

    new_readings = [r for r in readings if r["date"] not in existing_dates]
    new_purchases = [p for p in purchases
                     if (p["date"], p["supplier"].upper(), float(p["litres"] or 0)) not in existing_pkeys]
    new_distributions = [d for d in distributions
                         if (d["date"], d["store_location"], float(d["litres"] or 0)) not in existing_dkeys]

    print("\n=== INSERT PLAN ===")
    print(f"  New diesel_readings:      {len(new_readings)} (skip {len(readings)-len(new_readings)} dupes)")
    print(f"  New diesel_purchases:     {len(new_purchases)} (skip {len(purchases)-len(new_purchases)} dupes)")
    print(f"  New diesel_distributions: {len(new_distributions)} (skip {len(distributions)-len(new_distributions)} dupes)")

    if not apply_mode:
        print("\nDRY RUN — no writes performed. Re-run with --apply to write.")
        return

    print("\n=== WRITING ===")
    sn_to_id = {}
    if new_purchases:
        payload = [{
            "date": p["date"], "supplier": p["supplier"], "litres": p["litres"],
            "price_per_litre": p["price_per_litre"], "total_cost": p["total_cost"],
        } for p in new_purchases]
        inserted = sb.insert("diesel_purchases", payload)
        ins_map = {(x["date"], (x["supplier"] or "").upper(), float(x["litres"])): x["id"] for x in inserted}
        for p in new_purchases:
            key = (p["date"], p["supplier"].upper(), float(p["litres"] or 0))
            if key in ins_map:
                sn_to_id[p["_sn"]] = ins_map[key]
        print(f"  + Inserted {len(inserted)} purchases")

    # Refresh full purchase map for distribution linking
    all_p = sb.select("diesel_purchases", columns="id,date,supplier,litres")
    pkey_to_id = {(x["date"], (x["supplier"] or "").upper(), float(x["litres"])): x["id"] for x in all_p}
    for p in purchases:
        key = (p["date"], p["supplier"].upper(), float(p["litres"] or 0))
        if key in pkey_to_id:
            sn_to_id[p["_sn"]] = pkey_to_id[key]

    if new_distributions:
        payload = [{
            "date": d["date"], "store_location": d["store_location"], "litres": d["litres"],
            "purchase_id": sn_to_id.get(d["purchase_sn"]) if d["purchase_sn"] else None,
            "is_overage": d["is_overage"],
            "received_confirmed": False,
            "notes": f"Imported (raw: {d['raw_store']})" if d["is_overage"] else None,
        } for d in new_distributions]
        BATCH = 100; total_in = 0
        for i in range(0, len(payload), BATCH):
            chunk = payload[i:i+BATCH]
            inserted = sb.insert("diesel_distributions", chunk)
            total_in += len(inserted)
        print(f"  + Inserted {total_in} distributions")

    if new_readings:
        payload = []
        for r in new_readings:
            note_parts = []
            if r["supplier"]: note_parts.append(f"Supplier: {r['supplier']}")
            if r["received_by"]: note_parts.append(f"Transfer-out: {str(r['received_by'])[:200]}")
            payload.append({
                "generator_id": gen_id,
                "store_location": AKURE_STORE_NAME,
                "date": r["date"],
                "gen_hours_opening": r["gen_h_open"],
                "gen_hours_closing": r["gen_h_close"],
                "diesel_level_actual": r["closing_main"],
                "diesel_added": r["purchases_in"] or 0,
                "consumption_litres": int(r["consumption_l"]) if r["consumption_l"] is not None else None,
                "consumption_rate": r["consumption_rate"],
                "nepa_meter_opening": r["nepa_open"],
                "nepa_meter_closing": r["nepa_close"],
                "submitted_by": None,
                "notes": " | ".join(note_parts) if note_parts else None,
            })
        BATCH = 100; total_in = 0
        for i in range(0, len(payload), BATCH):
            chunk = payload[i:i+BATCH]
            inserted = sb.insert("diesel_readings", chunk)
            total_in += len(inserted)
        print(f"  + Inserted {total_in} diesel readings for Akure 1")

    print("\nDONE.")


if __name__ == "__main__":
    apply_mode = "--apply" in sys.argv
    main(apply_mode)
