#!/usr/bin/env python
r"""
FleetPro: Re-import ADO 1 daily diesel readings WITH transfer-out data.

Source: "ADO 1 DIESEL TRACKER (2).xlsx" — clean workbook, monthly tabs
JANUARY 2026..JULY 2026, location cell "ADO 1 CR". This replaces the original
Cowork-era import (which kept no transfer columns and left a phantom Apr-23
row) AND the store's sparse app entries.

COLUMN LAYOUT: the "narrow" template (same as Ondo 2 / Akungba JAN-FEB), date
in col 1. Verified all months: opening - closing + purchases - transfer_out =
consumption. Has a real NEPA meter (kWh counter, cols 16/17) and a TRANSFER
OUT column (col 6; only 2 rows, May 9-10, 180 L total, no vehicle named).

  1  DATE            2  SUPPLIER
  4  TOTAL OPENING   5  PURCHASES / TRANSFER IN  -> diesel_added
  6  DIESEL TRANSFER OUT -> diesel_transfers row (dest 'other') + note
  8  CLOSING MAIN    9  TOTAL CLOSING            -> diesel_level_actual
  10 CONSUMPTION     12/13 GEN RUNNING HOUR Open/Close
  14 HOUR RUN        15 rate (recomputed)         16/17 NEPA METER Open/Close

DATE-SHIFT GOTCHA (verified June): the store's APP entries carry yesterday's
meters under today's date (sheet Jun 22 3334.47->3336.07 == app "Jun 23"; sheet
Jul 1 3340.04->3340.24 == app "Jul 2"). The sheet dates the interval to the day
it ran, so in --replace mode the sheet wins on all its dates AND the app rows
in DUPLICATE_APP_DATES (same meter interval dated one day later, which would
otherwise double-count and run the meter backwards) are deleted as well.

JULY tab: only Jul 1 is real; Jul 2 is a trailing template artifact (blank
closing -> cons 1248) — excluded by requiring closing>0 OR gen-close>0.

After --apply, run the canonical post-import flow:
  py scripts\recalc_baselines.py --apply
  py scripts\backfill_discrepancy_flags.py --apply
  py scripts\sync_generator_hours.py --apply

Usage:
  py scripts\import_ado1.py                       # dry run
  py scripts\import_ado1.py --apply --replace     # sheet is authoritative

Reads SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY from .env / SupabaseCreds.env / env.
"""
import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
import datetime
from pathlib import Path

import openpyxl

FILE = Path(r"C:\Users\MakindeAribo\Downloads\ADO 1 DIESEL TRACKER (2).xlsx")
STORE_NAME = "Ado 1"
GEN_LOC_MATCH = "ado 1"
YEAR = 2026
# App rows whose meter interval the sheet already covers under the previous
# date (staff logged yesterday's meters under today's date).
DUPLICATE_APP_DATES = {"2026-07-02"}

MONTH_TABS = {
    "JANUARY 2026": 1, "FEBRUARY 2026": 2, "MARCH 2026": 3, "APRIL 2026": 4,
    "MAY 2026": 5, "JUNE 2026": 6, "JULY 2026": 7,
}

COL = {
    "supplier": 2, "opening_total": 4, "purchases_in": 5, "transfer_out": 6,
    "closing_main": 8, "closing_total": 9, "consumption_l": 10,
    "gen_h_open": 12, "gen_h_close": 13, "hour_run": 14, "consumption_rate": 15,
    "nepa_open": 16, "nepa_close": 17,
}


def load_env_file():
    for name in (".env", "SupabaseCreds.env", "supabase.env"):
        p = Path(name)
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
        return


class Supabase:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def _req(self, method, path, params=None, body=None, extra_headers=None):
        url = f"{self.url}/rest/v1/{path}"
        if params:
            url += "?" + urllib.parse.urlencode(params, doseq=True)
        data = json.dumps(body).encode("utf-8") if body is not None else None
        headers = dict(self.headers)
        if extra_headers:
            headers.update(extra_headers)
        req = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read()
                return json.loads(raw) if raw else []
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {e.read().decode('utf-8','replace')}") from e

    def select(self, table, columns="*", filters=None):
        params = {"select": columns}
        if filters:
            params.update(filters)
        return self._req("GET", table, params=params)

    def insert(self, table, rows):
        return self._req("POST", table, body=rows, extra_headers={"Prefer": "return=representation"})

    def delete(self, table, filters):
        return self._req("DELETE", table, params=filters, extra_headers={"Prefer": "return=minimal"})


def n(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return None if (isinstance(v, float) and v != v) else float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("#"):
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


def parse_readings():
    wb = openpyxl.load_workbook(FILE, data_only=True)
    out = []
    for tab, month in MONTH_TABS.items():
        if tab not in wb.sheetnames:
            print(f"  ! missing tab {tab!r}")
            continue
        ws = wb[tab]
        cnt = 0
        for r in range(1, ws.max_row + 1):
            d = ws.cell(r, 1).value
            if not isinstance(d, datetime.datetime) or d.year != YEAR or d.month != month:
                continue
            closing = n(ws.cell(r, COL["closing_total"]).value)
            if closing is None or closing <= 0:
                closing = n(ws.cell(r, COL["closing_main"]).value)
            gh_close = n(ws.cell(r, COL["gen_h_close"]).value)
            # Real day = closing level OR closing gen-hours; excludes the July
            # trailing template artifact (blank closing -> cons = opening).
            if not ((closing is not None and closing > 0) or (gh_close is not None and gh_close > 0)):
                continue
            gh_open = n(ws.cell(r, COL["gen_h_open"]).value)
            cons = n(ws.cell(r, COL["consumption_l"]).value)
            hours = (gh_close - gh_open) if (gh_open is not None and gh_close is not None) else None
            rate = round(cons / hours, 2) if (cons is not None and hours and hours > 0.05) else None
            tx = n(ws.cell(r, COL["transfer_out"]).value) or 0
            sup = ws.cell(r, COL["supplier"]).value
            sup = sup.strip() if isinstance(sup, str) and sup.strip() else None
            notes = []
            if sup:
                notes.append(f"Supplier: {sup}")
            if tx > 0:
                notes.append(f"Transfer-out: {tx:.0f}")
            out.append({
                "date": d.date().isoformat(),
                "gen_h_open": gh_open, "gen_h_close": gh_close,
                "closing": round(closing, 1) if closing is not None else None,
                "purchases_in": n(ws.cell(r, COL["purchases_in"]).value),
                "consumption_l": round(cons, 1) if cons is not None else None,
                "consumption_rate": rate,
                "nepa_open": n(ws.cell(r, COL["nepa_open"]).value),
                "nepa_close": n(ws.cell(r, COL["nepa_close"]).value),
                "transfer_out": tx,
                "notes": "; ".join(notes) if notes else None,
            })
            cnt += 1
        print(f"  {tab:<14}: {cnt} rows")
    out.sort(key=lambda r: r["date"])
    return out


def main(apply_mode, replace_mode):
    load_env_file()
    url, key = os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY required.")
        sys.exit(1)
    sb = Supabase(url, key)

    print("=== PARSING ADO 1 ===")
    readings = parse_readings()
    tot_tx = sum(r["transfer_out"] for r in readings)
    print(f"  Total: {len(readings)} rows | purchases {sum(r['purchases_in'] or 0 for r in readings):,.0f} L | transfers-out {tot_tx:,.0f} L\n")

    print("=== SUPABASE LOOKUP ===")
    gens = sb.select("generators", columns="id,name,loc,asset_type")
    gen = next((g for g in gens if (g.get("loc") or "").strip().lower() == GEN_LOC_MATCH
                and (g.get("asset_type") or "generator") != "oven"), None)
    if not gen:
        print(f"ERROR: no generator with loc == {STORE_NAME!r}.")
        sys.exit(1)
    gen_id = gen["id"]
    print(f"  Ado 1 generator: {gen['name']} (id={gen_id})")
    existing = sb.select("diesel_readings", columns="id,date,submitted_by,gen_hours_opening,gen_hours_closing",
                         filters={"generator_id": f"eq.{gen_id}"})
    print(f"  Existing readings: {len(existing)}")

    sheet_dates = {r["date"] for r in readings}
    overlap = [r for r in existing if r["date"] in sheet_dates or r["date"] in DUPLICATE_APP_DATES]

    if replace_mode:
        new_readings = readings
        print(f"\n=== REPLACE MODE ===")
        print(f"  Sheet wins on its dates + duplicate app dates {sorted(DUPLICATE_APP_DATES)}.")
        print(f"  Will DELETE {len(overlap)} existing reading(s), insert {len(readings)}.")
        app_over = [r for r in overlap if r.get("submitted_by")]
        if app_over:
            print(f"  ({len(app_over)} app-entered: {', '.join(sorted(r['date'] for r in app_over))})")
    else:
        existing_dates = {r["date"] for r in existing}
        new_readings = [r for r in readings if r["date"] not in existing_dates]

    print(f"\n=== PLAN ===")
    print(f"  Insert: {len(new_readings)} readings ({new_readings[0]['date']} .. {new_readings[-1]['date']})")
    tx_rows = [r for r in readings if r["transfer_out"] > 0]
    print(f"  diesel_transfers to upsert: {len(tx_rows)} ({tot_tx:.0f} L)")

    if not apply_mode:
        print("\nDRY RUN — no writes. Re-run with --apply --replace.")
        return

    print("\n=== WRITING ===")
    if replace_mode and overlap:
        for r in overlap:
            sb.delete("diesel_readings", {"id": f"eq.{r['id']}"})
        print(f"  - Deleted {len(overlap)} reading(s).")
    if new_readings:
        payload = [{
            "generator_id": gen_id, "store_location": STORE_NAME, "date": r["date"],
            "gen_hours_opening": r["gen_h_open"], "gen_hours_closing": r["gen_h_close"],
            "diesel_level_actual": r["closing"], "diesel_added": r["purchases_in"] or 0,
            "consumption_litres": r["consumption_l"],
            "consumption_rate": r["consumption_rate"],
            "nepa_meter_opening": r["nepa_open"], "nepa_meter_closing": r["nepa_close"],
            "submitted_by": None, "notes": r["notes"],
        } for r in new_readings]
        BATCH = 100
        total = 0
        for i in range(0, len(payload), BATCH):
            total += len(sb.insert("diesel_readings", payload[i:i + BATCH]))
        print(f"  + Inserted {total} Ado 1 readings")
    # Transfers: dedup by (date, store, litres) so re-runs are safe.
    existing_tx = sb.select("diesel_transfers", columns="date,store_location,litres",
                            filters={"store_location": f"eq.{STORE_NAME}"})
    have = {(t["date"], float(t["litres"])) for t in existing_tx}
    added_tx = 0
    for r in tx_rows:
        if (r["date"], float(r["transfer_out"])) in have:
            continue
        sb.insert("diesel_transfers", [{
            "date": r["date"], "store_location": STORE_NAME,
            "source_generator_id": gen_id, "dest_type": "other", "dest_id": None,
            "dest_label": "Transfer out (sheet)", "litres": r["transfer_out"],
            "notes": "From ADO 1 DIESEL TRACKER import", "recorded_by": None,
        }])
        added_tx += 1
    print(f"  + Inserted {added_tx} diesel_transfers rows")
    print("\nDONE. Next: recalc_baselines.py --apply -> backfill_discrepancy_flags.py --apply -> sync_generator_hours.py --apply")


if __name__ == "__main__":
    main("--apply" in sys.argv, "--replace" in sys.argv)
