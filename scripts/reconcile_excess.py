#!/usr/bin/env python
r"""
FleetPro: Reconcile historical diesel excess from the sheet's notes.

The STORES SUPPLY / SUMMARY tabs carry notes like:
  "AFTER SUPPLY, THERE WAS 600 LITRES OF EXCESS DIESEL OBTAINED FROM DIESEL
   BULK OF 33,000 LITRES"
Each note sits just below the bulk-tanker cycle that produced the excess, so
the bulk that generated it is the most recent big purchase (>= 30,000 L, or the
note's stated bulk size) ABOVE the note in the sheet.

This script walks the sheet, pairs each excess note with that bulk's DATE +
SIZE, then in the DB sets that purchase's litres_received = paid + excess
(matching on litres == bulk size and the nearest date). Result: Stock in Hand
(received - distributed) collapses toward zero and each big tanker shows its
documented excess.

Idempotent: re-running just re-sets the same litres_received values.

Usage:
  py scripts\reconcile_excess.py            # dry run (shows every pairing)
  py scripts\reconcile_excess.py --apply     # write litres_received

Run AFTER the 20260612_purchase_received_litres.sql migration.
"""
import json
import os
import re
import sys
import datetime
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl

FILE = Path(r"C:\Users\MakindeAribo\Downloads\DIESEL REPORT TEMPLATE (4).xlsx")
BIG = 30000  # litres at/above which a purchase counts as a bulk tanker


def load_env_file():
    for name in (".env", "SupabaseCreds.env", "supabase.env"):
        p = Path(name)
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
        return


class Supabase:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def _req(self, method, path, params=None, body=None, extra_headers=None):
        full = f"{self.url}/rest/v1/{path}"
        if params:
            full += "?" + urllib.parse.urlencode(params, doseq=True)
        data = json.dumps(body).encode("utf-8") if body is not None else None
        headers = dict(self.headers)
        if extra_headers:
            headers.update(extra_headers)
        req = urllib.request.Request(full, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read()
                return json.loads(raw) if raw else []
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {e.read().decode('utf-8','replace')}") from e

    def select_all(self, table, columns, page=1000):
        out = []
        offset = 0
        while True:
            chunk = self._req("GET", table, params={"select": columns, "limit": str(page), "offset": str(offset)})
            out.extend(chunk)
            if len(chunk) < page:
                return out
            offset += page

    def patch(self, table, id_val, body):
        return self._req("PATCH", table, params={"id": f"eq.{id_val}"}, body=body,
                         extra_headers={"Prefer": "return=minimal"})


def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def parse_date(v, last):
    if isinstance(v, datetime.datetime):
        d = v.date()
        if d.day <= 12 and d.month <= 12 and d.day != d.month and last:
            try:
                sw = datetime.date(d.year, d.day, d.month)
            except ValueError:
                return d
            best = min((d, sw), key=lambda c: (abs((c - last).days), c != d))
            return d if (best != d and abs((best - last).days) > 60) else best
        return d
    if isinstance(v, str):
        for fmt in ("%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.datetime.strptime(v.strip().replace("-", "/"), fmt).date()
            except ValueError:
                continue
    return None


EXCESS_RE = re.compile(r"(\d[\d,]*)\s*(?:LITRE|LITER|L)S?\s+(?:OF\s+)?EXCESS", re.I)
BULK_RE = re.compile(r"BULK\s+OF\s+(\d[\d,]*)", re.I)


def collect_notes(wb):
    """Walk STORES SUPPLY top-to-bottom; for each excess note find the most
    recent big purchase (date + size) above it."""
    ws = wb["STORES SUPPLY"]
    last_date = None
    last_bulk = None   # (date, litres) of most recent big purchase
    pairings = []
    for r in range(1, ws.max_row + 1):
        d = parse_date(ws.cell(r, 1).value, last_date)
        if d:
            last_date = d
        bought = num(ws.cell(r, 3).value)
        if bought and bought >= BIG and last_date:
            last_bulk = (last_date.isoformat(), bought)
        # excess note may be in any cell of the row
        for c in range(1, min(ws.max_column + 1, 14)):
            v = ws.cell(r, c).value
            if not isinstance(v, str) or "EXCESS" not in v.upper():
                continue
            flat = " ".join(v.split())
            em = EXCESS_RE.search(flat)
            bm = BULK_RE.search(flat)
            if not em:
                continue  # blank excess ("THERE WAS LITRES EXCESS") - skip
            excess = num(em.group(1))
            bulk_size = num(bm.group(1)) if bm else (last_bulk[1] if last_bulk else None)
            if not excess or not bulk_size:
                continue
            bulk_date = last_bulk[0] if last_bulk else None
            pairings.append({"row": r, "excess": excess, "bulk_size": bulk_size, "bulk_date": bulk_date})
            break
    return pairings


def main(apply_mode):
    load_env_file()
    url, key = os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY required.")
        sys.exit(1)
    sb = Supabase(url, key)
    wb = openpyxl.load_workbook(FILE, data_only=True)

    pairings = collect_notes(wb)
    print(f"Excess notes found (with amount): {len(pairings)}\n")

    purchases = sb.select_all("diesel_purchases", "id,date,supplier,litres,litres_received")
    used = set()  # purchase ids already matched, so two notes don't grab the same tanker

    print(f"{'note@row':<9} {'bulk date':<12} {'bulk':>7} {'excess':>7}  -> matched purchase")
    updates = []
    unmatched = []
    for pr in sorted(pairings, key=lambda x: x["bulk_date"] or ""):
        # candidates: same litres as the bulk, not already used
        cands = [p for p in purchases if abs(float(p["litres"]) - pr["bulk_size"]) < 1 and p["id"] not in used]
        if not cands:
            unmatched.append(pr)
            print(f"R{pr['row']:<7} {pr['bulk_date'] or '?':<12} {pr['bulk_size']:>7.0f} {pr['excess']:>7.0f}  -> (no unused {pr['bulk_size']:.0f} L purchase)")
            continue
        # A note describes the bulk that was just distributed, so it was bought
        # on/before the note. Prefer the LATEST unused tanker on-or-before the
        # note date; only if none exists fall back to the earliest one after.
        if pr["bulk_date"]:
            nd = datetime.date.fromisoformat(pr["bulk_date"])
            before = sorted([p for p in cands if datetime.date.fromisoformat(p["date"]) <= nd],
                            key=lambda p: p["date"], reverse=True)
            after = sorted([p for p in cands if datetime.date.fromisoformat(p["date"]) > nd],
                           key=lambda p: p["date"])
            cands = before + after
        chosen = cands[0]
        used.add(chosen["id"])
        new_recv = float(chosen["litres"]) + pr["excess"]
        updates.append((chosen["id"], new_recv, chosen))
        print(f"R{pr['row']:<7} {pr['bulk_date'] or '?':<12} {pr['bulk_size']:>7.0f} {pr['excess']:>7.0f}  -> {chosen['date']} {chosen['supplier']} (recv {new_recv:.0f})")

    print(f"\n=== SUMMARY ===")
    print(f"  Notes matched: {len(updates)}   unmatched: {len(unmatched)}")
    print(f"  Total excess attributed: {sum(u[1]-float(u[2]['litres']) for u in updates):,.0f} L")

    if not apply_mode:
        print("\nDRY RUN — no writes. Re-run with --apply to set litres_received.")
        return

    print("\n=== WRITING ===")
    for pid, recv, _ in updates:
        sb.patch("diesel_purchases", pid, {"litres_received": recv})
    print(f"  Set litres_received on {len(updates)} purchases.")
    print("\nDONE. Stock in Hand (received - distributed) should now be near zero.")
    print("If you re-allocate links: py scripts\\relink_purchase_ids.py --apply")


if __name__ == "__main__":
    main("--apply" in sys.argv)
