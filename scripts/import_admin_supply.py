#!/usr/bin/env python
r"""
FleetPro: Import/refresh the admin diesel supply log from DIESEL REPORT TEMPLATE (3).

Tabs handled:
  DIESEL PURCHASE   -> diesel_purchases   (bulk buys; dates with DD/MM swaps fixed)
  STORES SUPPLY     -> diesel_distributions (admin -> store deliveries, Mar 2025 - now)
  OUTSOURCED SUPPLY -> diesel_purchases (no price recorded -> cost null)
                       + diesel_distributions (direct-to-store, notes 'Outsourced: X')

Dedup:
  purchases     by (date, litres)        - survives supplier renames in the sheet
                                           (e.g. BOVAS -> BOVAS (AKURE))
  distributions by (date, store, litres) - re-listing the same delivery in two
                                           tabs collapses to one row

Store-name normalization handles all spelling variants; ALAGBAKA* maps to
Akure 1 (per owner), OKITIPUPA (bare) is assumed Okitipupa CR.

Usage:
  py scripts\import_admin_supply.py            # dry run
  py scripts\import_admin_supply.py --apply     # write
"""
import json
import os
import re
import sys
import datetime
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl

FILE = Path(r"C:\Users\MakindeAribo\Downloads\DIESEL REPORT TEMPLATE (4).xlsx")

# One-off corrections to purchases already in the DB: (old_date, litres) -> new_date.
# SHARFA 2,000 L was bought 29 Apr (per STORES SUPPLY col C); the purchase tab had a 29/05 typo.
DATE_CORRECTIONS = {("2026-05-29", 2000.0): "2026-04-29"}

# Tanker-size litres accepted for UNNAMED col-C purchase rows (only with --include-unnamed)
TANKER_SIZES = {33000.0, 33980.0, 35000.0}

# ============================================================
# VERIFIED_TANKERS — fill this in as you confirm real purchases.
#
# The sheet has unnamed tanker-size rows; some are real purchases, some are
# echoes of purchases already imported. When you confirm one is REAL (see the
# balance-jump check in the dry-run instructions), add its date here with the
# supplier name (use "UNKNOWN" if you don't know who supplied it):
#
#   VERIFIED_TANKERS = {
#       "2025-03-03": "BOVAS",
#       "2025-05-05": "UNKNOWN",
#   }
#
# Then re-run:  py scripts\import_admin_supply.py --apply
#               py scripts\relink_purchase_ids.py --apply
# ============================================================
VERIFIED_TANKERS = {}

STORE_MAP = {
    "ADO 1": "Ado 1", "ADO 2": "Ado 2", "ADO 3": "Ado 3", "ADO BAKERY": "Ado Bakery",
    "AKUNGBA": "Akungba CR", "AKUNGBA C.R.": "Akungba CR", "AKUNGBA C.R": "Akungba CR", "AKUNGBA CR": "Akungba CR",
    "AKURE 1": "Akure 1", "AKURE 2": "Akure 2", "AKURE 3": "Akure 3", "AKURE 4": "Akure 4",
    "AKURE 5": "Akure 5", "AKURE 6": "Akure 6", "AKURE BAKERY": "Akure Bakery",
    "ALAGBAKA": "Akure 1", "ALAGBAKA GEN": "Akure 1", "ALAGBAKA GEN.": "Akure 1",
    "ALAGABAKA GEN": "Akure 1", "ALAGBAKA GEN(OVERAGE)": "Akure 1", "OVERAGE - AKURE 1": "Akure 1",
    "IDANRE": "Idanre CR", "IGBOKODA": "Igbokoda CR", "IGBOKODA CR": "Igbokoda CR",
    "IKARE BAKERY": "Ikare Bakery", "IKARE BAKEKRY": "Ikare Bakery", "IKARE BAKKERY": "Ikare Bakery",
    "IKARE C.R.": "Ikare CR", "IKARE C.R": "Ikare CR", "IKARE CR": "Ikare CR",
    "OKITIPUPA": "Okitipupa CR",  # bare name: assumed CR (1 row)
    "OKITIPUPA BAKERY": "Okitipupa Bakery",
    "OKITIPUPA C.R": "Okitipupa CR", "OKITIPUPA C.R.": "Okitipupa CR", "OKITIPUPA CR": "Okitipupa CR",
    "OND0 1": "Ondo CR", "ONDO 1": "Ondo CR", "ONDO CR": "Ondo CR",
    "ONDO 2": "Ondo 2 CR", "ONDO CR 2": "Ondo 2 CR", "ONDO 2 CR": "Ondo 2 CR",
    "ONDO BAKERY": "Ondo Bakery",
    "OWO BAKERY": "Owo Bakery",
    "OWO C.R.": "Owo CR", "OWO C.R": "Owo CR", "OWO CR": "Owo CR",
    "OYE BAKERY": "Oye Bakery",
    "PIE EXPRESS/WAREHOUSE": "Pie Express / Warehouse",
}
OVERAGE_SOURCES = {"ALAGBAKA GEN(OVERAGE)", "OVERAGE - AKURE 1"}


def load_env_file():
    for name in (".env", "SupabaseCreds.env", "supabase.env"):
        p = Path(name)
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
        return


class Supabase:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def _req(self, method, path, params=None, body=None, extra_headers=None):
        full = f"{self.url}/rest/v1/{path}"
        if params:
            full += "?" + urllib.parse.urlencode(params, doseq=True)
        data = json.dumps(body).encode("utf-8") if body is not None else None
        headers = dict(self.headers)
        if extra_headers:
            headers.update(extra_headers)
        req = urllib.request.Request(full, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read()
                return json.loads(raw) if raw else []
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {e.read().decode('utf-8','replace')}") from e

    def select_all(self, table, columns, filters=None, page=1000):
        out = []
        offset = 0
        while True:
            params = {"select": columns, "limit": str(page), "offset": str(offset)}
            if filters:
                params.update(filters)
            chunk = self._req("GET", table, params=params)
            out.extend(chunk)
            if len(chunk) < page:
                return out
            offset += page

    def insert(self, table, rows):
        return self._req("POST", table, body=rows, extra_headers={"Prefer": "return=representation"})

    def patch(self, table, id_val, body):
        return self._req("PATCH", table, params={"id": f"eq.{id_val}"}, body=body,
                         extra_headers={"Prefer": "return=representation"})


def num(v):
    if isinstance(v, (int, float)):
        return None if (isinstance(v, float) and v != v) else float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(",", "").strip())
        except ValueError:
            return None
    return None


def parse_date_cell(v, last_date):
    """Date cell may be a datetime (possibly day/month-swapped by Excel) or a
    'DD/MM/YYYY' string. For ambiguous datetimes, pick the variant that best
    continues the chronology."""
    if isinstance(v, datetime.datetime):
        d = v.date()
        if d.day <= 12 and d.month <= 12 and d.day != d.month and last_date:
            try:
                swapped = datetime.date(d.year, d.day, d.month)
            except ValueError:
                return d
            # Pick whichever candidate sits closest to the running chronology
            # (ties favour the raw value), but never swap across a >60-day jump —
            # out-of-order re-listed blocks keep their raw date instead.
            best = min((d, swapped), key=lambda c: (abs((c - last_date).days), c != d))
            if best != d and abs((best - last_date).days) > 60:
                return d
            return best
        return d
    if isinstance(v, str):
        s = v.strip().replace("-", "/")
        for fmt in ("%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def parse_purchases_bulk(wb):
    ws = wb["DIESEL PURCHASE"]
    out = []
    last = None
    for r in range(2, ws.max_row + 1):
        sn = ws.cell(r, 1).value
        if sn == "TOTAL" or not isinstance(sn, int):
            continue
        raw_date = ws.cell(r, 3).value
        if raw_date is None:
            continue  # OPENING row
        d = parse_date_cell(raw_date, last)
        if d is None:
            print(f"  ! bulk purchase row {r}: unparseable date {raw_date!r}")
            continue
        last = d
        litres = num(ws.cell(r, 4).value) or 0
        supplier = (ws.cell(r, 5).value or "").strip() if isinstance(ws.cell(r, 5).value, str) else ""
        price = num(ws.cell(r, 6).value)
        out.append({"date": d.isoformat(), "supplier": supplier, "litres": litres, "price_per_litre": price})
    return out


def parse_stores_supply(wb):
    """Returns (distributions, unmapped, colc_named, colc_unnamed).

    Col C of STORES SUPPLY is the warehouse's real purchase log, but it is
    polluted with running totals and block echoes. Rows with a textual
    supplier in col D are trusted purchases; unnamed rows are returned
    separately and only imported when they look like standard tanker loads
    AND the user passes --include-unnamed."""
    ws = wb["STORES SUPPLY"]
    today = datetime.date.today().isoformat()
    out = []
    colc_named = []
    colc_unnamed = []
    last_date = None
    last_supplier = None
    unmapped = {}
    for r in range(3, ws.max_row + 1):
        d = parse_date_cell(ws.cell(r, 1).value, last_date)
        if d:
            last_date = d
        sup_raw = ws.cell(r, 4).value
        sup_is_name = isinstance(sup_raw, str) and any(ch.isalpha() for ch in sup_raw)
        if sup_is_name:
            last_supplier = sup_raw.strip()
        # Col C: purchases into the warehouse
        bought = num(ws.cell(r, 3).value)
        if bought and bought > 0 and last_date:
            if sup_is_name:
                colc_named.append({"date": last_date.isoformat(), "supplier": sup_raw.strip(),
                                   "litres": bought, "price_per_litre": None})
            elif bought in TANKER_SIZES and last_date.isoformat() <= today:
                colc_unnamed.append({"date": last_date.isoformat(), "supplier": "UNKNOWN (tanker)",
                                     "litres": bought, "price_per_litre": None})
        qty = num(ws.cell(r, 5).value)
        store_raw = ws.cell(r, 6).value
        if not qty or qty <= 0 or not store_raw or not last_date:
            continue
        key = str(store_raw).strip().upper()
        if key in {"BALANCE", "STORE SUPPLIED"}:
            continue
        mapped = STORE_MAP.get(key)
        if not mapped:
            unmapped[str(store_raw).strip()] = unmapped.get(str(store_raw).strip(), 0) + 1
            continue
        out.append({"date": last_date.isoformat(), "store": mapped, "litres": qty,
                    "supplier": last_supplier, "is_overage": key in OVERAGE_SOURCES, "src": "supply"})
    return out, unmapped, colc_named, colc_unnamed


def parse_outsourced(wb):
    ws = wb["OUTSOURCED SUPPLY"]
    purchases = []
    dists = []
    unmapped = {}
    last_date = None
    last_supplier = None
    for r in range(3, ws.max_row + 1):
        d = parse_date_cell(ws.cell(r, 1).value, last_date)
        if d:
            last_date = d
        sup = ws.cell(r, 2).value
        if sup and isinstance(sup, str) and sup.strip():
            last_supplier = sup.strip()
        bought = num(ws.cell(r, 3).value)
        qty = num(ws.cell(r, 4).value)
        store_raw = ws.cell(r, 5).value
        if bought and bought > 0 and last_date and last_supplier:
            purchases.append({"date": last_date.isoformat(), "supplier": last_supplier,
                              "litres": bought, "price_per_litre": None})
        if qty and qty > 0 and store_raw and last_date:
            key = str(store_raw).strip().upper()
            if key in {"BALANCE", "STORE SUPPLIED"}:
                continue
            mapped = STORE_MAP.get(key)
            if not mapped:
                unmapped[str(store_raw).strip()] = unmapped.get(str(store_raw).strip(), 0) + 1
                continue
            dists.append({"date": last_date.isoformat(), "store": mapped, "litres": qty,
                          "supplier": last_supplier, "is_overage": key in OVERAGE_SOURCES, "src": "outsourced"})
    return purchases, dists, unmapped


def main(apply_mode, include_unnamed):
    load_env_file()
    url, key = os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY required.")
        sys.exit(1)
    sb = Supabase(url, key)
    wb = openpyxl.load_workbook(FILE, data_only=True)

    print("=== PARSING ===")
    bulk = parse_purchases_bulk(wb)
    print(f"  DIESEL PURCHASE: {len(bulk)} rows")
    supply, un1, colc_named, colc_unnamed = parse_stores_supply(wb)
    print(f"  STORES SUPPLY:   {len(supply)} distributions, col-C purchases: {len(colc_named)} named, {len(colc_unnamed)} unnamed tanker-size")
    out_p, out_d, un2 = parse_outsourced(wb)
    print(f"  OUTSOURCED:      {len(out_p)} purchases, {len(out_d)} distributions")
    for label, un in (("STORES SUPPLY", un1), ("OUTSOURCED", un2)):
        for name, cnt in sorted(un.items()):
            print(f"  ! unmapped store in {label}: {name!r} x{cnt} (skipped)")

    # ---- Date corrections ----
    # Apply to the PARSED sheet rows too, otherwise the purchase tab's typo'd
    # row re-imports as a fresh purchase right after we correct the DB one.
    for plist in (bulk, out_p, colc_named, colc_unnamed):
        for p in plist:
            k = (p["date"], float(p["litres"]))
            if k in DATE_CORRECTIONS:
                p["date"] = DATE_CORRECTIONS[k]
    db_purch = sb.select_all("diesel_purchases", "id,date,supplier,litres,price_per_litre")
    for (old_date, litres), new_date in DATE_CORRECTIONS.items():
        hit = next((p for p in db_purch if p["date"] == old_date and float(p["litres"]) == litres), None)
        if hit:
            print(f"  CORRECTION: {hit['supplier']} {litres:.0f} L  {old_date} -> {new_date}" + ("" if apply_mode else "  (dry run)"))
            if apply_mode:
                sb.patch("diesel_purchases", hit["id"], {"date": new_date})
            hit["date"] = new_date  # in-memory too, so the dedup plan reflects the correction

    # ---- Merge purchase sources ----
    # Conflict guard: a named col-C row matching an existing purchase on
    # (date, supplier-prefix) but with DIFFERENT litres is probably the same
    # event recorded inconsistently between tabs — skip it and warn.
    existing_ds = {(p["date"], (p.get("supplier") or "").upper()[:5]) for p in db_purch}
    colc_ok = []
    for p in colc_named:
        k = (p["date"], p["supplier"].upper()[:5])
        if k in existing_ds and (p["date"], float(p["litres"])) not in {(x["date"], float(x["litres"])) for x in db_purch}:
            print(f"  ! CONFLICT (skipped): col-C {p['date']} {p['supplier']} {p['litres']:.0f} L disagrees with an existing purchase same day/supplier — reconcile manually")
            continue
        colc_ok.append(p)
    # Verified unnamed tankers (dates listed in VERIFIED_TANKERS) are imported
    # with their confirmed supplier; the rest stay out unless --include-unnamed.
    verified = []
    still_unnamed = []
    for p in colc_unnamed:
        if p["date"] in VERIFIED_TANKERS:
            p["supplier"] = VERIFIED_TANKERS[p["date"]]
            verified.append(p)
        else:
            still_unnamed.append(p)
    if verified:
        print(f"\n  VERIFIED tankers to import ({len(verified)}):")
        for p in sorted(verified, key=lambda x: x["date"]):
            print(f"    {p['date']}  {p['litres']:>7.0f} L  {p['supplier']}")
    colc_ok = colc_ok + verified
    colc_unnamed = still_unnamed
    if colc_unnamed and not include_unnamed:
        print(f"\n  UNNAMED tanker-size rows NOT imported ({len(colc_unnamed)}) — verify each (does the sheet's warehouse balance jump that day?),")
        print(f"  then add real ones to VERIFIED_TANKERS at the top of this script:")
        for p in sorted(colc_unnamed, key=lambda x: x["date"]):
            print(f"    {p['date']}  {p['litres']:>7.0f} L")

    pkeys = {(p["date"], float(p["litres"])) for p in db_purch}
    db_dist = sb.select_all("diesel_distributions", "date,store_location,litres")
    dkeys = {(x["date"], x["store_location"], float(x["litres"])) for x in db_dist}

    all_purch = bulk + out_p + colc_ok + (colc_unnamed if include_unnamed else [])
    new_purch = []
    seen_p = set(pkeys)
    for p in all_purch:
        k = (p["date"], float(p["litres"]))
        if k in seen_p:
            continue
        seen_p.add(k)
        new_purch.append(p)

    all_dist = supply + out_d
    new_dist = []
    seen_d = set(dkeys)
    in_sheet_dupes = 0
    for x in all_dist:
        k = (x["date"], x["store"], float(x["litres"]))
        if k in seen_d:
            if k not in dkeys:
                in_sheet_dupes += 1
            continue
        seen_d.add(k)
        new_dist.append(x)

    print("\n=== INSERT PLAN ===")
    print(f"  Purchases:     {len(new_purch)} new (skip {len(all_purch)-len(new_purch)} dupes)")
    for p in new_purch:
        price = f"@{p['price_per_litre']:.0f}" if p["price_per_litre"] else "(no price)"
        print(f"    {p['date']}  {p['litres']:>7.0f} L  {p['supplier']:<18} {price}")
    print(f"  Distributions: {len(new_dist)} new (skip {len(all_dist)-len(new_dist)} dupes, {in_sheet_dupes} within-sheet)")
    by_store = {}
    for x in new_dist:
        s = by_store.setdefault(x["store"], {"n": 0, "litres": 0.0})
        s["n"] += 1
        s["litres"] += x["litres"]
    for store, s in sorted(by_store.items()):
        print(f"    {store:<26} {s['n']:>4} rows  {s['litres']:>9.0f} L")
    if new_dist:
        print(f"    date range: {min(x['date'] for x in new_dist)} .. {max(x['date'] for x in new_dist)}")

    if not apply_mode:
        print("\nDRY RUN — no writes. Re-run with --apply to write.")
        return

    print("\n=== WRITING ===")
    if new_purch:
        # price_per_litre is NOT NULL in the schema — unpriced (outsourced) buys
        # are stored as 0; app-side averages exclude zero-priced litres.
        payload = [{"date": p["date"], "supplier": p["supplier"], "litres": p["litres"],
                    "price_per_litre": p["price_per_litre"] or 0} for p in new_purch]
        ins = sb.insert("diesel_purchases", payload)
        print(f"  + {len(ins)} purchases")

    # Refresh purchase map for linking
    db_purch = sb.select_all("diesel_purchases", "id,date,supplier,litres")

    def link(dist):
        sup = (dist.get("supplier") or "").upper()
        if not sup:
            return None
        cands = [p for p in db_purch if (p.get("supplier") or "").upper().startswith(sup[:5]) and p["date"] <= dist["date"]]
        return max(cands, key=lambda p: p["date"])["id"] if cands else None

    if new_dist:
        payload = [{"date": x["date"], "store_location": x["store"], "litres": x["litres"],
                    "purchase_id": link(x), "is_overage": x["is_overage"], "received_confirmed": False,
                    "notes": ("Outsourced: " + (x.get("supplier") or "")) if x["src"] == "outsourced" else None}
                   for x in new_dist]
        BATCH = 100
        total = 0
        for i in range(0, len(payload), BATCH):
            total += len(sb.insert("diesel_distributions", payload[i:i + BATCH]))
        print(f"  + {total} distributions")
    print("\nDONE.")


if __name__ == "__main__":
    main("--apply" in sys.argv, "--include-unnamed" in sys.argv)
