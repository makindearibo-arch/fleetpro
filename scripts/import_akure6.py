#!/usr/bin/env python
r"""
FleetPro: Re-import AKURE 6 daily diesel readings WITH transfer columns.

Source: "AK6 Daily Diesel Tracker New (1).xlsx" - 17 monthly tabs Mar 2025 ..
Jul 2026 (quirky names: "OCT. 2025", "JAN2026", "MAY,2026", "SEPT 2025").
Replaces the Cowork-era import (467 rows, no transfer data, 49% flagged).

Narrow layout (same as Ado 1 / Ondo CR), date col 1:
  2 SUPPLIER | 4 TOTAL OPENING | 5 PURCHASES/IN -> diesel_added
  6 TRANSFER OUT -> diesel_transfers (bare litres) | 8/9 CLOSING (use 9,
  fallback 8) -> diesel_level_actual | 10 CONSUMPTION | 12/13 GEN HOURS
  Open/Close | 14 HOUR RUN | 15 rate | 16/17 NEPA

DATA-QUALITY HANDLING: Akure 6's meter discipline is poor (some cells hold a
DATE instead of an hour value, e.g. Mar-2025 close "1934-11-16"; backwards
meters). num() returns None for any non-numeric cell, so a corrupted meter just
drops that row's hours (its level data still imports). Rows are collected by
their REAL date value (not the tab name, since the "MARCH 2025" tab starts on
Feb 20) and de-duplicated date-wise, later tab winning.

Usage:
  py scripts\import_akure6.py                       # dry run
  py scripts\import_akure6.py --apply --replace     # sheet is authoritative

Then: recalc_baselines.py --apply -> backfill_discrepancy_flags.py --apply ->
sync_generator_hours.py --apply
"""
import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
import datetime
from pathlib import Path

import openpyxl

FILE = Path(r"C:\Users\MakindeAribo\Downloads\AK6 Daily Diesel Tracker New (1).xlsx")
STORE_NAME = "Akure 6"
GEN_LOC_MATCH = "akure 6"

MONTHLY_TABS = [
    "MARCH 2025", "APRIL 2025", "MAY 2025", "JUNE 2025", "JULY 2025", "AUGUST 2025",
    "SEPT 2025", "OCT. 2025", "NOV 2025", "DEC 2025", "JAN2026", "FEB 2026",
    "MARCH 2026", "APRIL 2026", "MAY,2026", "JUNE 2026", "JULY 2026",
]

COL = {
    "supplier": 2, "opening_total": 4, "purchases_in": 5, "transfer_out": 6,
    "closing_main": 8, "closing_total": 9, "consumption_l": 10,
    "gen_h_open": 12, "gen_h_close": 13, "hour_run": 14, "consumption_rate": 15,
    "nepa_open": 16, "nepa_close": 17,
}


def load_env_file():
    for name in (".env", "SupabaseCreds.env", "supabase.env"):
        p = Path(name)
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
        return


class Supabase:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def _req(self, method, path, params=None, body=None, extra_headers=None):
        url = f"{self.url}/rest/v1/{path}"
        if params:
            url += "?" + urllib.parse.urlencode(params, doseq=True)
        data = json.dumps(body).encode("utf-8") if body is not None else None
        headers = dict(self.headers)
        if extra_headers:
            headers.update(extra_headers)
        req = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read()
                return json.loads(raw) if raw else []
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {e.read().decode('utf-8','replace')}") from e

    def select(self, table, columns="*", filters=None):
        params = {"select": columns}
        if filters:
            params.update(filters)
        return self._req("GET", table, params=params)

    def insert(self, table, rows):
        return self._req("POST", table, body=rows, extra_headers={"Prefer": "return=representation"})

    def delete(self, table, filters):
        return self._req("DELETE", table, params=filters, extra_headers={"Prefer": "return=minimal"})


def n(v):
    if v is None or isinstance(v, datetime.datetime):
        return None
    if isinstance(v, (int, float)):
        return None if (isinstance(v, float) and v != v) else float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("#"):
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


def parse_readings():
    wb = openpyxl.load_workbook(FILE, data_only=True)
    by_date = {}
    for tab in MONTHLY_TABS:
        if tab not in wb.sheetnames:
            print(f"  ! missing tab {tab!r}")
            continue
        ws = wb[tab]
        cnt = 0
        for r in range(1, ws.max_row + 1):
            d = ws.cell(r, 1).value
            if not isinstance(d, datetime.datetime):
                continue
            closing = n(ws.cell(r, COL["closing_total"]).value)
            if closing is None or closing <= 0:
                closing = n(ws.cell(r, COL["closing_main"]).value)
            gh_close = n(ws.cell(r, COL["gen_h_close"]).value)
            if not ((closing is not None and closing > 0) or (gh_close is not None and gh_close > 0)):
                continue
            gh_open = n(ws.cell(r, COL["gen_h_open"]).value)
            cons = n(ws.cell(r, COL["consumption_l"]).value)
            hours = (gh_close - gh_open) if (gh_open is not None and gh_close is not None) else None
            rate = round(cons / hours, 2) if (cons is not None and hours and hours > 0.05) else None
            tx = n(ws.cell(r, COL["transfer_out"]).value) or 0
            sup = ws.cell(r, COL["supplier"]).value
            sup = sup.strip() if isinstance(sup, str) and sup.strip() else None
            notes = []
            if sup:
                notes.append(f"Supplier: {sup}")
            if tx > 0:
                notes.append(f"Transfer-out: {tx:.0f}")
            by_date[d.date().isoformat()] = {
                "date": d.date().isoformat(),
                "gen_h_open": gh_open, "gen_h_close": gh_close,
                "closing": round(closing, 1) if closing is not None else None,
                "purchases_in": n(ws.cell(r, COL["purchases_in"]).value),
                "consumption_l": round(cons, 1) if cons is not None else None,
                "consumption_rate": rate,
                "nepa_open": n(ws.cell(r, COL["nepa_open"]).value),
                "nepa_close": n(ws.cell(r, COL["nepa_close"]).value),
                "transfer_out": tx,
                "notes": "; ".join(notes) if notes else None,
            }
            cnt += 1
        print(f"  {tab:<12}: {cnt} rows")
    out = list(by_date.values())
    out.sort(key=lambda r: r["date"])
    return out


def main(apply_mode, replace_mode):
    load_env_file()
    url, key = os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY required.")
        sys.exit(1)
    sb = Supabase(url, key)

    print("=== PARSING AKURE 6 ===")
    readings = parse_readings()
    tot_tx = sum(r["transfer_out"] for r in readings)
    print(f"  Total: {len(readings)} rows ({readings[0]['date']}..{readings[-1]['date']}) | "
          f"purchases {sum(r['purchases_in'] or 0 for r in readings):,.0f} L | transfers-out {tot_tx:,.0f} L\n")

    print("=== SUPABASE LOOKUP ===")
    gens = sb.select("generators", columns="id,name,loc,asset_type")
    gen = next((g for g in gens if (g.get("loc") or "").strip().lower() == GEN_LOC_MATCH
                and (g.get("asset_type") or "generator") != "oven"), None)
    if not gen:
        print(f"ERROR: no generator with loc == {STORE_NAME!r}.")
        sys.exit(1)
    gen_id = gen["id"]
    print(f"  Akure 6 generator: {gen['name']} (id={gen_id})")
    existing = sb.select("diesel_readings", columns="id,date,submitted_by",
                         filters={"generator_id": f"eq.{gen_id}"})
    print(f"  Existing readings: {len(existing)}")

    sheet_dates = {r["date"] for r in readings}
    # This sheet is the complete authoritative history, so --replace clears ALL
    # existing rows (incl. ~31 stale Cowork-era rows before the sheet's Feb-20
    # start), not just overlapping dates.
    overlap = existing if replace_mode else [r for r in existing if r["date"] in sheet_dates]

    if replace_mode:
        new_readings = readings
        print(f"\n=== REPLACE MODE ===")
        print(f"  Will DELETE ALL {len(overlap)} existing reading(s), insert {len(readings)}.")
        app_over = [r for r in overlap if r.get("submitted_by")]
        if app_over:
            print(f"  ({len(app_over)} app-entered: {', '.join(sorted(r['date'] for r in app_over))})")
    else:
        existing_dates = {r["date"] for r in existing}
        new_readings = [r for r in readings if r["date"] not in existing_dates]

    print(f"\n=== PLAN ===")
    print(f"  Insert: {len(new_readings)} readings")
    tx_rows = [r for r in readings if r["transfer_out"] > 0]
    print(f"  diesel_transfers to upsert: {len(tx_rows)} ({tot_tx:.0f} L)")

    if not apply_mode:
        print("\nDRY RUN — no writes. Re-run with --apply --replace.")
        return

    print("\n=== WRITING ===")
    if replace_mode and overlap:
        for i in range(0, len(overlap), 1):
            sb.delete("diesel_readings", {"id": f"eq.{overlap[i]['id']}"})
        print(f"  - Deleted {len(overlap)} reading(s).")
    if new_readings:
        payload = [{
            "generator_id": gen_id, "store_location": STORE_NAME, "date": r["date"],
            "gen_hours_opening": r["gen_h_open"], "gen_hours_closing": r["gen_h_close"],
            "diesel_level_actual": r["closing"], "diesel_added": r["purchases_in"] or 0,
            "consumption_litres": r["consumption_l"], "consumption_rate": r["consumption_rate"],
            "nepa_meter_opening": r["nepa_open"], "nepa_meter_closing": r["nepa_close"],
            "submitted_by": None, "notes": r["notes"],
        } for r in new_readings]
        BATCH = 100
        total = 0
        for i in range(0, len(payload), BATCH):
            total += len(sb.insert("diesel_readings", payload[i:i + BATCH]))
        print(f"  + Inserted {total} Akure 6 readings")
    existing_tx = sb.select("diesel_transfers", columns="date,litres",
                            filters={"store_location": f"eq.{STORE_NAME}"})
    have = {(t["date"], float(t["litres"])) for t in existing_tx}
    added_tx = 0
    for r in tx_rows:
        if (r["date"], float(r["transfer_out"])) in have:
            continue
        sb.insert("diesel_transfers", [{
            "date": r["date"], "store_location": STORE_NAME,
            "source_generator_id": gen_id, "dest_type": "other", "dest_id": None,
            "dest_label": "Transfer out (sheet)", "litres": r["transfer_out"],
            "notes": "From AK6 tracker import", "recorded_by": None,
        }])
        added_tx += 1
    print(f"  + Inserted {added_tx} diesel_transfers rows")
    print("\nDONE. Next: recalc_baselines.py --apply -> backfill_discrepancy_flags.py --apply -> sync_generator_hours.py --apply")


if __name__ == "__main__":
    main("--apply" in sys.argv, "--replace" in sys.argv)
