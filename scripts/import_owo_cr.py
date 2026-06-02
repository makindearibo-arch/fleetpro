#!/usr/bin/env python
r"""
FleetPro: Import OWO CR daily diesel readings.

NOTE ON COLUMN LAYOUT: The OWO CR template differs from the Akure 1 template.
OWO CR has NO "DIESEL RECEIVED BY" column, so every column from G (7) rightward
is shifted LEFT by one relative to Akure 1. The COLMAP below reflects the OWO
layout exactly (verified: opening - closing = consumption; close - open = hours run).

Purchases / distributions are NOT touched here — the supply template is unchanged
from the last import, so there is nothing new on that side.

Usage:
  py scripts\import_owo_cr.py                      # dry run
  py scripts\import_owo_cr.py --apply              # write readings
  py scripts\import_owo_cr.py --apply --recalc-baseline

Reads SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY from .env / SupabaseCreds.env / env.
"""
import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
import datetime
from pathlib import Path

import openpyxl

OWO_FILE = Path(r"C:\Users\MakindeAribo\Downloads\OWO CR DIESEL TEMPLATE 2026.xlsx")
STORE_NAME = "Owo CR"          # must match generators.loc / locations.name
GEN_LOC_MATCH = "owo cr"        # case-insensitive loc match to find the generator

MONTHLY_TABS = {
    "JANUARY 2026": (2026, 1), "FEBRUARY 2026": (2026, 2), "MARCH 2026": (2026, 3),
    "APRIL 2026": (2026, 4), "MAY 2026": (2026, 5), "JUNE 2026": (2026, 6),
}

# OWO CR column map (1-indexed). Verified against header rows + arithmetic.
COL = {
    "date": 1,
    "supplier": 2,
    "opening_main": 3,
    "purchases_in": 5,      # PURCHASES / TRANSFER IN
    "transfer_out": 6,      # DIESEL TRANSFER OUT (to vehicles) — informational
    "available": 7,
    "closing_main": 8,      # CLOSING STOCK MAIN TANK -> diesel_level_actual
    "closing_total": 9,
    "consumption_l": 10,    # CONSUMPTION IN LITERS
    "gen_h_open": 12,       # GENERATOR RUNNING HOUR Opening
    "gen_h_close": 13,      # GENERATOR RUNNING HOUR Closing
    "consumption_rate": 15, # GENERATOR CONSUMPTION (LITRES/HOUR)
    "nepa_open": 16,        # NEPA METER Opening
    "nepa_close": 17,       # NEPA Closing
}


# ---------- .env ----------
def load_env_file():
    for name in (".env", "SupabaseCreds.env", "supabase.env"):
        p = Path(name)
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
        return


# ---------- Supabase REST ----------
class Supabase:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def _req(self, method, path, params=None, body=None, extra_headers=None):
        url = f"{self.url}/rest/v1/{path}"
        if params:
            url += "?" + urllib.parse.urlencode(params, doseq=True)
        data = json.dumps(body).encode("utf-8") if body is not None else None
        headers = dict(self.headers)
        if extra_headers:
            headers.update(extra_headers)
        req = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read()
                return json.loads(raw) if raw else []
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {e.read().decode('utf-8','replace')}") from e

    def select(self, table, columns="*", filters=None):
        params = {"select": columns}
        if filters:
            params.update(filters)
        return self._req("GET", table, params=params)

    def insert(self, table, rows):
        return self._req("POST", table, body=rows, extra_headers={"Prefer": "return=representation"})


def n(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return None if (isinstance(v, float) and v != v) else float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("#"):
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


def parse_readings():
    wb = openpyxl.load_workbook(OWO_FILE, data_only=True)
    out = []
    for tab, (year, month) in MONTHLY_TABS.items():
        if tab not in wb.sheetnames:
            print(f"  ! missing tab {tab!r}")
            continue
        ws = wb[tab]
        start = None
        for r in range(1, ws.max_row + 1):
            if isinstance(ws.cell(r, 1).value, datetime.datetime):
                start = r
                break
        if start is None:
            print(f"  ! no data rows in {tab}")
            continue
        cnt = 0
        for r in range(start, ws.max_row + 1):
            d = ws.cell(r, COL["date"]).value
            if not isinstance(d, datetime.datetime) or d.year != year or d.month != month:
                continue
            rec = {
                "date": d.date().isoformat(),
                "gen_h_open": n(ws.cell(r, COL["gen_h_open"]).value),
                "gen_h_close": n(ws.cell(r, COL["gen_h_close"]).value),
                "closing_main": n(ws.cell(r, COL["closing_main"]).value),
                "purchases_in": n(ws.cell(r, COL["purchases_in"]).value),
                "consumption_l": n(ws.cell(r, COL["consumption_l"]).value),
                "consumption_rate": n(ws.cell(r, COL["consumption_rate"]).value),
                "nepa_open": n(ws.cell(r, COL["nepa_open"]).value),
                "nepa_close": n(ws.cell(r, COL["nepa_close"]).value),
                "supplier": (ws.cell(r, COL["supplier"]).value or "").strip() if isinstance(ws.cell(r, COL["supplier"]).value, str) else None,
            }
            if rec["gen_h_open"] is None and rec["gen_h_close"] is None and rec["closing_main"] is None:
                continue
            out.append(rec)
            cnt += 1
        print(f"  {tab}: {cnt} rows")
    return out


def recalc_baseline(sb, gen_id):
    rows = sb._req("GET", "diesel_readings", params={
        "select": "date,gen_hours_opening,gen_hours_closing,diesel_level_actual,diesel_added",
        "generator_id": f"eq.{gen_id}", "order": "date.asc",
    })
    rates = []
    prev = None
    for r in rows:
        ho, hc, lvl = r.get("gen_hours_opening"), r.get("gen_hours_closing"), r.get("diesel_level_actual")
        added = r.get("diesel_added") or 0
        hrs = (hc - ho) if (ho is not None and hc is not None) else None
        if hrs and hrs > 0 and lvl is not None and prev is not None:
            actual = prev + added - lvl
            if actual > 0:
                rate = actual / hrs
                if 1 <= rate <= 100:
                    rates.append(rate)
        if lvl is not None:
            prev = lvl
    if not rates:
        print("  No valid rate pairs — skipping baseline.")
        return
    avg = sum(rates) / len(rates)
    flagged = sum(1 for r in rates if abs(r - avg) / avg > 0.20) / len(rates) * 100
    print(f"  Pairs: {len(rates)}  Avg: {avg:.2f} L/hr (min {min(rates):.2f}, max {max(rates):.2f})  ~{flagged:.1f}% >20% off mean")
    sb._req("POST", "generator_baselines", params={"on_conflict": "generator_id"},
            body=[{"generator_id": gen_id, "avg_litres_per_hour": round(avg, 2),
                   "baseline_readings_count": len(rates),
                   "last_calculated": datetime.datetime.now(datetime.timezone.utc).isoformat(),
                   "min_rate": round(min(rates), 2), "max_rate": round(max(rates), 2)}],
            extra_headers={"Prefer": "resolution=merge-duplicates,return=representation"})
    print("  Baseline upserted.")


def main(apply_mode, recalc_mode):
    load_env_file()
    url, key = os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY required.")
        sys.exit(1)
    sb = Supabase(url, key)

    print("=== PARSING OWO CR ===")
    readings = parse_readings()
    print(f"  Total: {len(readings)} reading rows\n")

    print("=== SUPABASE LOOKUP ===")
    gens = sb.select("generators", columns="id,name,loc")
    gen = next((g for g in gens if (g.get("loc") or "").strip().lower() == GEN_LOC_MATCH), None)
    if not gen:
        print(f"ERROR: no generator with loc='{STORE_NAME}'. Found locs:")
        for g in gens:
            print(f"  - {g['name']!r} loc={g.get('loc')!r}")
        sys.exit(1)
    gen_id = gen["id"]
    print(f"  OWO CR generator: {gen['name']} (id={gen_id}, loc={gen.get('loc')})")
    existing = sb.select("diesel_readings", columns="date", filters={"generator_id": f"eq.{gen_id}"})
    existing_dates = {r["date"] for r in existing}
    print(f"  Existing readings for this generator: {len(existing_dates)}")

    new_readings = [r for r in readings if r["date"] not in existing_dates]
    print(f"\n=== INSERT PLAN ===")
    print(f"  New diesel_readings: {len(new_readings)} (skip {len(readings) - len(new_readings)} dupes)")
    if new_readings:
        print(f"  Date range: {new_readings[0]['date']} .. {new_readings[-1]['date']}")

    if not apply_mode:
        print("\nDRY RUN — no writes. Re-run with --apply to write.")
        if recalc_mode:
            print("\n=== BASELINE RECALC (OWO CR) ===")
            recalc_baseline(sb, gen_id)
        return

    print("\n=== WRITING ===")
    if new_readings:
        payload = [{
            "generator_id": gen_id, "store_location": STORE_NAME, "date": r["date"],
            "gen_hours_opening": r["gen_h_open"], "gen_hours_closing": r["gen_h_close"],
            "diesel_level_actual": r["closing_main"], "diesel_added": r["purchases_in"] or 0,
            "consumption_litres": int(r["consumption_l"]) if r["consumption_l"] is not None else None,
            "consumption_rate": r["consumption_rate"],
            "nepa_meter_opening": r["nepa_open"], "nepa_meter_closing": r["nepa_close"],
            "submitted_by": None,
            "notes": f"Supplier: {r['supplier']}" if r["supplier"] else None,
        } for r in new_readings]
        BATCH = 100
        total = 0
        for i in range(0, len(payload), BATCH):
            total += len(sb.insert("diesel_readings", payload[i:i + BATCH]))
        print(f"  + Inserted {total} OWO CR readings")
    print("\nDONE.")
    if recalc_mode:
        print("\n=== BASELINE RECALC (OWO CR) ===")
        recalc_baseline(sb, gen_id)


if __name__ == "__main__":
    main("--apply" in sys.argv, "--recalc-baseline" in sys.argv)
