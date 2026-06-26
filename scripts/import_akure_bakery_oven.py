#!/usr/bin/env python
r"""
FleetPro: Update AKURE BAKERY oven daily diesel readings from the latest file.

Source: "Daily Diesel Tracker Bakery Oven AKURE BAKERY 2026 (1).xlsx" — monthly
tabs JANUARY 2026..JUNE2026, location cell "AKURE BAKERY". This is an OVEN
(asset_type='oven'): it tracks BATCHES PRODUCED and L/batch, NOT hour meter or
NEPA. Matches the bakery template used by scripts/import_bakeries.py:
  col 2  SUPPLIER
  col 4  TOTAL OPENING STOCK
  col 5  PURCHASES / TRANSFER IN  -> diesel_added
  col 8  CLOSING STOCK - MAIN     -> diesel_level_actual (col 9 fallback)
  col 10 CONSUMPTION IN LITRES
  col 12 BATCHES PRODUCED         -> batches_produced
  col 13 OVEN CONSUMPTION (L/batch) -> consumption_rate
Date column is auto-detected per tab (1 or 2) and all columns anchor off it.

--replace makes the sheet authoritative for any date it covers: it deletes
existing readings on those dates (incl. sparse app-entered ones) and inserts
clean rows. Ovens have no hour-meter baseline, so no baseline recalc.

Usage:
  py scripts\import_akure_bakery_oven.py                  # dry run
  py scripts\import_akure_bakery_oven.py --apply --replace

Reads SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY from .env / SupabaseCreds.env / env.
"""
import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
import datetime
from pathlib import Path

import openpyxl

FILE = Path(r"C:\Users\MakindeAribo\Downloads\Daily Diesel Tracker Bakery Oven AKURE BAKERY 2026 (1).xlsx")
STORE_NAME = "Akure Bakery"
GEN_LOC_MATCH = "akure bakery"
YEAR = 2026

MONTH_TABS = {
    "JANUARY 2026": 1, "FEBRUARY 2026": 2, "MARCH 2026": 3,
    "APRIL2026": 4, "MAY2026": 5, "JUNE2026": 6,
}

# Offsets from the detected date column (base = date_col - 1).
OFF = {
    "supplier": 2, "opening_total": 4, "purchases_in": 5, "transfer_out": 6,
    "closing_main": 8, "closing_total": 9, "consumption_l": 10,
    "batches": 12, "rate": 13,
}


def load_env_file():
    for name in (".env", "SupabaseCreds.env", "supabase.env"):
        p = Path(name)
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
        return


class Supabase:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def _req(self, method, path, params=None, body=None, extra_headers=None):
        url = f"{self.url}/rest/v1/{path}"
        if params:
            url += "?" + urllib.parse.urlencode(params, doseq=True)
        data = json.dumps(body).encode("utf-8") if body is not None else None
        headers = dict(self.headers)
        if extra_headers:
            headers.update(extra_headers)
        req = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read()
                return json.loads(raw) if raw else []
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {e.read().decode('utf-8','replace')}") from e

    def select(self, table, columns="*", filters=None):
        params = {"select": columns}
        if filters:
            params.update(filters)
        return self._req("GET", table, params=params)

    def insert(self, table, rows):
        return self._req("POST", table, body=rows, extra_headers={"Prefer": "return=representation"})

    def delete(self, table, filters):
        return self._req("DELETE", table, params=filters, extra_headers={"Prefer": "return=minimal"})


def n(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return None if (isinstance(v, float) and v != v) else float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("#"):
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


def detect_date_col(ws):
    for dc in (1, 2):
        for r in range(1, ws.max_row + 1):
            if isinstance(ws.cell(r, dc).value, datetime.datetime):
                return dc
    return None


def parse_readings():
    wb = openpyxl.load_workbook(FILE, data_only=True)
    out = []
    for tab, month in MONTH_TABS.items():
        if tab not in wb.sheetnames:
            print(f"  ! missing tab {tab!r}")
            continue
        ws = wb[tab]
        dc = detect_date_col(ws)
        if dc is None:
            continue
        base = dc - 1
        def col(key):
            return base + OFF[key]
        cnt = 0
        for r in range(1, ws.max_row + 1):
            d = ws.cell(r, dc).value
            if not isinstance(d, datetime.datetime) or d.year != YEAR or d.month != month:
                continue
            closing = n(ws.cell(r, col("closing_main")).value)
            if closing is None or closing <= 0:
                closing = n(ws.cell(r, col("closing_total")).value)
            batches = n(ws.cell(r, col("batches")).value)
            cons = n(ws.cell(r, col("consumption_l")).value)
            # A real oven day has a closing tank level OR batches produced. A
            # trailing template row (blank closing -> 0, no batches) leaves a
            # bogus consumption = opening - 0; exclude it (consumption alone is
            # not enough to count a row as real).
            if not ((closing is not None and closing > 0) or (batches or 0) > 0):
                continue
            rate = n(ws.cell(r, col("rate")).value)
            if rate is None and cons is not None and batches and batches > 0:
                rate = round(cons / batches, 2)
            sup = ws.cell(r, col("supplier")).value
            sup = sup.strip() if isinstance(sup, str) and sup.strip() else None
            out.append({
                "date": d.date().isoformat(),
                "closing": closing,
                "purchases_in": n(ws.cell(r, col("purchases_in")).value),
                "consumption_l": cons,
                "rate": rate,
                "batches": int(batches) if batches is not None else None,
                "supplier": sup,
            })
            cnt += 1
        print(f"  {tab:<14} datecol={dc}: {cnt} rows")
    out.sort(key=lambda r: r["date"])
    return out


def main(apply_mode, replace_mode):
    load_env_file()
    url, key = os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY required.")
        sys.exit(1)
    sb = Supabase(url, key)

    print("=== PARSING AKURE BAKERY OVEN ===")
    readings = parse_readings()
    print(f"  Total: {len(readings)} reading rows  ({sum(r['batches'] or 0 for r in readings):,} batches)\n")

    print("=== SUPABASE LOOKUP ===")
    gens = sb.select("generators", columns="id,name,loc,asset_type")
    gen = next((g for g in gens if (g.get("loc") or "").strip().lower() == GEN_LOC_MATCH
                and (g.get("asset_type") or "") == "oven"), None)
    if not gen:
        gen = next((g for g in gens if GEN_LOC_MATCH in (g.get("loc") or "").strip().lower()
                    and (g.get("asset_type") or "") == "oven"), None)
    if not gen:
        print(f"ERROR: no OVEN with loc ~ {STORE_NAME!r}. Akure Bakery assets:")
        for g in gens:
            if GEN_LOC_MATCH in (g.get("loc") or "").lower():
                print(f"  - {g['name']!r} loc={g.get('loc')!r} asset_type={g.get('asset_type')!r}")
        sys.exit(1)
    gen_id = gen["id"]
    print(f"  Akure Bakery oven: {gen['name']} (id={gen_id}, loc={gen.get('loc')})")
    existing = sb.select("diesel_readings", columns="id,date,submitted_by", filters={"generator_id": f"eq.{gen_id}"})
    existing_dates = {r["date"] for r in existing}
    print(f"  Existing readings for this oven: {len(existing_dates)}")

    sheet_dates = {r["date"] for r in readings}
    overlap = [r for r in existing if r["date"] in sheet_dates]

    if replace_mode:
        new_readings = readings
        print(f"\n=== REPLACE MODE ===")
        print(f"  The sheet wins on overlapping dates. Will DELETE {len(overlap)} existing reading(s), then insert all {len(readings)}.")
        app_over = [r for r in overlap if r.get("submitted_by")]
        if app_over:
            print(f"  ({len(app_over)} of those are app-entered: {', '.join(sorted(r['date'] for r in app_over))})")
    else:
        new_readings = [r for r in readings if r["date"] not in existing_dates]

    print(f"\n=== PLAN ===")
    print(f"  diesel_readings to insert: {len(new_readings)} "
          f"({'replace mode — all sheet rows' if replace_mode else f'skip {len(readings) - len(new_readings)} dupes'})")
    if new_readings:
        print(f"  Date range: {new_readings[0]['date']} .. {new_readings[-1]['date']}")
        print(f"  Diesel added total: {sum(r['purchases_in'] or 0 for r in new_readings):,.0f} L")

    if not apply_mode:
        print("\nDRY RUN — no writes. Re-run with --apply --replace.")
        return

    print("\n=== WRITING ===")
    if replace_mode and overlap:
        for r in overlap:
            sb.delete("diesel_readings", {"id": f"eq.{r['id']}"})
        print(f"  - Deleted {len(overlap)} overlapping reading(s).")
    if new_readings:
        payload = [{
            "generator_id": gen_id, "store_location": STORE_NAME, "date": r["date"],
            "diesel_level_actual": r["closing"], "diesel_added": r["purchases_in"] or 0,
            "consumption_litres": int(r["consumption_l"]) if r["consumption_l"] is not None else None,
            "consumption_rate": r["rate"], "batches_produced": r["batches"],
            "submitted_by": None,
            "notes": f"Supplier: {r['supplier']}" if r["supplier"] else None,
        } for r in new_readings]
        BATCH = 100
        total = 0
        for i in range(0, len(payload), BATCH):
            total += len(sb.insert("diesel_readings", payload[i:i + BATCH]))
        print(f"  + Inserted {total} Akure Bakery oven readings")
    print("\nDONE.")


if __name__ == "__main__":
    main("--apply" in sys.argv, "--replace" in sys.argv)
