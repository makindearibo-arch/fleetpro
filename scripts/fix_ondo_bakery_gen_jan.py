#!/usr/bin/env python
r"""
FleetPro: Fix Ondo Bakery Generator's phantom JANUARY 2026 consumption.

Root cause: in "Generator Daily Diesel Tracker ondo bakery  (1).xlsx", the JAN
tab's closing-stock columns (8/9) were never filled (col 9 literally 0), so the
sheet's consumption formula = opening - 0 (+ purchases) = the WHOLE TANK every
day. That imported as 62,218 L for January alone (e.g. Jan 4: 760 + 2,000
purchase - 0 = "2,760 consumed") — the source of Ondo Bakery's impossible
71,673 L total. February onward the closings are filled and sane.

Fix: the real closings are the NEXT day's opening (col 4); Jan 31 closes at
FEB 1's opening. Real January consumption = opening + purchases - closing
(no transfers in Jan). Total comes to ~1,968 L, in line with other months.

This script deletes the 31 imported January 2026 readings for the Ondo Bakery
generator and reinserts them with derived closings/consumption. Idempotent.

Usage:
  py scripts\fix_ondo_bakery_gen_jan.py            # dry run
  py scripts\fix_ondo_bakery_gen_jan.py --apply
Then: recalc_baselines.py --apply -> backfill_discrepancy_flags.py --apply
"""
import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
import datetime
from pathlib import Path

import openpyxl

FILE = Path(r"C:\Users\MakindeAribo\Downloads\Generator Daily Diesel Tracker ondo bakery  (1).xlsx")
JAN_TAB = "JAN 2026 DAILY DIESEL TRACKER"
FEB_TAB = "feb 2026"
GEN_ID = "317d8e0d-2bb5-425e-862c-a9a729975ce2"   # Ondo Bakery Generator
STORE = "Ondo Bakery"


def load_env_file():
    for name in (".env", "SupabaseCreds.env", "supabase.env"):
        p = Path(name)
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
        return


class Supabase:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def _req(self, method, path, params=None, body=None, extra_headers=None):
        url = f"{self.url}/rest/v1/{path}"
        if params:
            url += "?" + urllib.parse.urlencode(params, doseq=True)
        data = json.dumps(body).encode("utf-8") if body is not None else None
        headers = dict(self.headers)
        if extra_headers:
            headers.update(extra_headers)
        req = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read()
                return json.loads(raw) if raw else []
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {e.read().decode('utf-8','replace')}") from e


def n(v):
    if isinstance(v, (int, float)):
        return float(v)
    return None


def main(apply_mode):
    load_env_file()
    url, key = os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY required.")
        sys.exit(1)
    sb = Supabase(url, key)
    wb = openpyxl.load_workbook(FILE, data_only=True)

    # Parse January rows: opening, purchases, hours, NEPA
    ws = wb[JAN_TAB]
    days = []
    for r in range(1, ws.max_row + 1):
        d = ws.cell(r, 1).value
        if not isinstance(d, datetime.datetime) or d.year != 2026 or d.month != 1:
            continue
        days.append({
            "date": d.date().isoformat(),
            "opening": n(ws.cell(r, 4).value),
            "pur": n(ws.cell(r, 5).value) or 0,
            "gh_open": n(ws.cell(r, 12).value),
            "gh_close": n(ws.cell(r, 13).value),
            "nepa_open": n(ws.cell(r, 16).value),
            "nepa_close": n(ws.cell(r, 17).value),
        })
    days.sort(key=lambda x: x["date"])
    # Feb 1 opening closes out Jan 31
    wsf = wb[FEB_TAB]
    feb1_open = None
    for r in range(1, wsf.max_row + 1):
        d = wsf.cell(r, 1).value
        if isinstance(d, datetime.datetime) and d.year == 2026 and d.month == 2 and d.day == 1:
            feb1_open = n(wsf.cell(r, 4).value)
            break
    print(f"January rows: {len(days)} | Feb 1 opening (= Jan 31 closing): {feb1_open}")

    new_rows = []
    tot = 0
    for i, d in enumerate(days):
        closing = days[i + 1]["opening"] if i + 1 < len(days) else feb1_open
        cons = None
        if d["opening"] is not None and closing is not None:
            cons = round(d["opening"] + d["pur"] - closing, 1)
            if cons < 0:
                cons = 0  # opening typo; don't record negative burn
        hrs = (d["gh_close"] - d["gh_open"]) if (d["gh_open"] is not None and d["gh_close"] is not None) else None
        rate = round(cons / hrs, 2) if (cons and hrs and hrs > 0) else None
        tot += cons or 0
        new_rows.append({
            "generator_id": GEN_ID, "store_location": STORE, "date": d["date"],
            "gen_hours_opening": d["gh_open"], "gen_hours_closing": d["gh_close"],
            "diesel_level_actual": closing, "diesel_added": d["pur"],
            "consumption_litres": cons, "consumption_rate": rate,
            "nepa_meter_opening": d["nepa_open"], "nepa_meter_closing": d["nepa_close"],
            "submitted_by": None,
            "notes": "Closing derived from next day's opening (sheet closing column blank in Jan)",
        })
    print(f"Derived January consumption: {tot:,.0f} L (was 62,218 phantom)")

    if not apply_mode:
        for x in new_rows[:5]:
            print(" ", x["date"], "close", x["diesel_level_actual"], "cons", x["consumption_litres"], "hrs",
                  (x["gen_hours_closing"] or 0) - (x["gen_hours_opening"] or 0))
        print("\nDRY RUN — no writes. Re-run with --apply.")
        return

    sb._req("DELETE", "diesel_readings",
            params={"generator_id": f"eq.{GEN_ID}", "date": "gte.2026-01-01", "and": "(date.lte.2026-01-31)"},
            extra_headers={"Prefer": "return=minimal"})
    print("Deleted old January rows.")
    out = sb._req("POST", "diesel_readings", body=new_rows, extra_headers={"Prefer": "return=representation"})
    print(f"Inserted {len(out)} corrected January readings.")
    print("\nDONE. Now run: recalc_baselines.py --apply -> backfill_discrepancy_flags.py --apply")


if __name__ == "__main__":
    main("--apply" in sys.argv)
