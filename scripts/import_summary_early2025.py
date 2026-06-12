#!/usr/bin/env python
r"""
FleetPro: Import the missing Jan-Feb 2025 records from the SUMMARY tab.

STORES SUPPLY barely covers January-February 2025, but the SUMMARY tab has
the weekly reports for that period: store deliveries (DISTRIBUTION / STORE
SUPPLIED columns) and one bulk purchase that exists nowhere else
(24 Jan 2025: 32,750 L BOVAS). This is the ~46,000 L "hole" that made
Stock in Hand overshoot and per-purchase attribution lag by one tanker.

Rules:
  - Only rows BEFORE 1 Mar 2025 are read (STORES SUPPLY covers from March).
  - Dedup by (date, store, litres) against existing distributions.
  - If an existing delivery has the same (date, store) but DIFFERENT litres,
    the SUMMARY row is skipped with a warning (conflicting amounts between
    tabs - e.g. Oye Bakery 16 Jan: outsourced tab says 2,000, SUMMARY says
    4,000 - reconcile manually rather than double-count).
  - Purchases likewise dedup by (date, litres) with a same-day/supplier
    conflict guard.

Usage:
  py scripts\import_summary_early2025.py            # dry run
  py scripts\import_summary_early2025.py --apply     # write
Then: py scripts\relink_purchase_ids.py --apply
"""
import json
import os
import sys
import datetime
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl

FILE = Path(r"C:\Users\MakindeAribo\Downloads\DIESEL REPORT TEMPLATE (4).xlsx")
CUTOFF = "2025-03-01"   # STORES SUPPLY takes over from here

STORE_MAP = {
    "ADO 1": "Ado 1", "ADO 2": "Ado 2", "ADO 3": "Ado 3", "ADO BAKERY": "Ado Bakery",
    "AKUNGBA": "Akungba CR", "AKUNGBA C.R.": "Akungba CR", "AKUNGBA C.R": "Akungba CR", "AKUNGBA CR": "Akungba CR",
    "AKURE 1": "Akure 1", "AKURE 2": "Akure 2", "AKURE 3": "Akure 3", "AKURE 4": "Akure 4",
    "AKURE 5": "Akure 5", "AKURE 6": "Akure 6", "AKURE BAKERY": "Akure Bakery",
    "ALAGBAKA": "Akure 1", "ALAGBAKA GEN": "Akure 1", "ALAGBAKA GEN.": "Akure 1",
    "ALAGABAKA GEN": "Akure 1", "ALAGBAKA GEN(OVERAGE)": "Akure 1", "OVERAGE - AKURE 1": "Akure 1",
    "IDANRE": "Idanre CR", "IGBOKODA": "Igbokoda CR", "IGBOKODA CR": "Igbokoda CR",
    "IKARE BAKERY": "Ikare Bakery", "IKARE BAKEKRY": "Ikare Bakery", "IKARE BAKKERY": "Ikare Bakery",
    "IKARE C.R.": "Ikare CR", "IKARE C.R": "Ikare CR", "IKARE CR": "Ikare CR",
    "OKITIPUPA": "Okitipupa CR", "OKITIPUPA BAKERY": "Okitipupa Bakery",
    "OKITIPUPA C.R": "Okitipupa CR", "OKITIPUPA C.R.": "Okitipupa CR", "OKITIPUPA CR": "Okitipupa CR",
    "OND0 1": "Ondo CR", "ONDO 1": "Ondo CR", "ONDO CR": "Ondo CR",
    "ONDO 2": "Ondo 2 CR", "ONDO CR 2": "Ondo 2 CR", "ONDO 2 CR": "Ondo 2 CR",
    "ONDO BAKERY": "Ondo Bakery",
    "OWO BAKERY": "Owo Bakery",
    "OWO C.R.": "Owo CR", "OWO C.R": "Owo CR", "OWO CR": "Owo CR",
    "OYE BAKERY": "Oye Bakery",
    "PIE EXPRESS/WAREHOUSE": "Pie Express / Warehouse",
}


def load_env_file():
    for name in (".env", "SupabaseCreds.env", "supabase.env"):
        p = Path(name)
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
        return


class Supabase:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def _req(self, method, path, params=None, body=None, extra_headers=None):
        full = f"{self.url}/rest/v1/{path}"
        if params:
            full += "?" + urllib.parse.urlencode(params, doseq=True)
        data = json.dumps(body).encode("utf-8") if body is not None else None
        headers = dict(self.headers)
        if extra_headers:
            headers.update(extra_headers)
        req = urllib.request.Request(full, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read()
                return json.loads(raw) if raw else []
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {e.read().decode('utf-8','replace')}") from e

    def select_all(self, table, columns, filters=None, page=1000):
        out = []
        offset = 0
        while True:
            params = {"select": columns, "limit": str(page), "offset": str(offset)}
            if filters:
                params.update(filters)
            chunk = self._req("GET", table, params=params)
            out.extend(chunk)
            if len(chunk) < page:
                return out
            offset += page

    def insert(self, table, rows):
        return self._req("POST", table, body=rows, extra_headers={"Prefer": "return=representation"})


def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s or s.upper() == "NIL":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def parse_date(v, last):
    if isinstance(v, datetime.datetime):
        d = v.date()
        if d.day <= 12 and d.month <= 12 and d.day != d.month and last:
            try:
                swapped = datetime.date(d.year, d.day, d.month)
            except ValueError:
                return d
            best = min((d, swapped), key=lambda c: (abs((c - last).days), c != d))
            if best != d and abs((best - last).days) > 60:
                return d
            return best
        return d
    if isinstance(v, str):
        s = v.strip().replace("-", "/")
        for fmt in ("%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def main(apply_mode):
    load_env_file()
    url, key = os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY required.")
        sys.exit(1)
    sb = Supabase(url, key)
    wb = openpyxl.load_workbook(FILE, data_only=True)
    ws = wb["SUMMARY"]

    last = None
    dists = []
    purchases = []
    unmapped = {}
    for r in range(3, ws.max_row + 1):
        d = parse_date(ws.cell(r, 1).value, last)
        if d:
            if d.isoformat() >= CUTOFF:
                break
            last = d
        if not last:
            continue
        # Purchase columns: RECEIVED(3) SUPPLIER(4) PRICE(6)
        rec = num(ws.cell(r, 3).value)
        sup = ws.cell(r, 4).value
        price = num(ws.cell(r, 6).value)
        if rec and rec > 0 and isinstance(sup, str) and any(c.isalpha() for c in sup):
            purchases.append({"date": last.isoformat(), "supplier": sup.strip(),
                              "litres": rec, "price_per_litre": price})
        # Distribution columns: DISTRIBUTION(8) STORE SUPPLIED(9)
        qty = num(ws.cell(r, 8).value)
        store_raw = ws.cell(r, 9).value
        if qty and qty > 0 and isinstance(store_raw, str) and store_raw.strip() and store_raw.strip().upper() != "NIL":
            keyname = store_raw.strip().upper()
            mapped = STORE_MAP.get(keyname)
            if not mapped:
                unmapped[store_raw.strip()] = unmapped.get(store_raw.strip(), 0) + 1
                continue
            dists.append({"date": last.isoformat(), "store": mapped, "litres": qty})

    print(f"=== PARSED (SUMMARY, before {CUTOFF}) ===")
    print(f"  Purchases: {len(purchases)} | Distributions: {len(dists)}")
    for n_, c in sorted(unmapped.items()):
        print(f"  ! unmapped store: {n_!r} x{c}")

    # ---- Dedup / conflict guards ----
    db_p = sb.select_all("diesel_purchases", "date,supplier,litres")
    pkeys = {(p["date"], float(p["litres"])) for p in db_p}
    p_ds = {(p["date"], (p.get("supplier") or "").upper()[:5]) for p in db_p}
    db_d = sb.select_all("diesel_distributions", "date,store_location,litres")
    dkeys = {(x["date"], x["store_location"], float(x["litres"])) for x in db_d}
    d_dstore = {(x["date"], x["store_location"]) for x in db_d}

    new_p = []
    seen = set(pkeys)
    for p in purchases:
        k = (p["date"], float(p["litres"]))
        if k in seen:
            continue
        if (p["date"], p["supplier"].upper()[:5]) in p_ds:
            print(f"  ! purchase conflict (skipped): {p['date']} {p['supplier']} {p['litres']:.0f} L — same day/supplier exists with different litres")
            continue
        seen.add(k)
        new_p.append(p)

    new_d = []
    seend = set(dkeys)
    for x in dists:
        k = (x["date"], x["store"], float(x["litres"]))
        if k in seend:
            continue
        if (x["date"], x["store"]) in d_dstore:
            print(f"  ! delivery conflict (skipped): {x['date']} {x['store']} {x['litres']:.0f} L — same day/store exists with different litres")
            continue
        seend.add(k)
        new_d.append(x)

    print(f"\n=== INSERT PLAN ===")
    print(f"  Purchases: {len(new_p)} new")
    for p in new_p:
        pr = f"@{p['price_per_litre']:.0f}" if p["price_per_litre"] else "(no price)"
        print(f"    {p['date']}  {p['litres']:>7.0f} L  {p['supplier']:<16} {pr}")
    print(f"  Distributions: {len(new_d)} new, {sum(x['litres'] for x in new_d):,.0f} L")
    by_store = {}
    for x in new_d:
        by_store[x["store"]] = by_store.get(x["store"], 0) + x["litres"]
    for s, l in sorted(by_store.items()):
        print(f"    {s:<26} {l:>8,.0f} L")

    if not apply_mode:
        print("\nDRY RUN — no writes. Re-run with --apply to write.")
        return

    print("\n=== WRITING ===")
    if new_p:
        ins = sb.insert("diesel_purchases", [{"date": p["date"], "supplier": p["supplier"],
                                              "litres": p["litres"],
                                              "price_per_litre": p["price_per_litre"] or 0} for p in new_p])
        print(f"  + {len(ins)} purchases")
    if new_d:
        ins = sb.insert("diesel_distributions", [{"date": x["date"], "store_location": x["store"],
                                                  "litres": x["litres"], "purchase_id": None,
                                                  "is_overage": False, "received_confirmed": False,
                                                  "notes": "From SUMMARY tab (Jan-Feb 2025)"} for x in new_d])
        print(f"  + {len(ins)} distributions")
    print("\nDONE. Now run: py scripts\\relink_purchase_ids.py --apply")


if __name__ == "__main__":
    main("--apply" in sys.argv)
