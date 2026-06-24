#!/usr/bin/env python
r"""
FleetPro: Import IKARE CR daily diesel readings.

Source: "IKARE CR DIESEL TEMPLATE YEAR 2026.xlsx". The workbook holds many
unrelated trackers; the diesel data is in the monthly tabs JAN..JUNE 2026.
Location cell reads "IKARE CR".

COLUMN LAYOUT (verified: opening - closing + purchases = consumption;
close - open = hours run; NEPA close - open = usage). This template puts the
DIESEL COST at col 12 and the GENERATOR RUNNING HOUR at cols 13/14 — i.e.
everything from cost rightward is one column further right than the Akure 2
template. The closing tank level is filled in BOTH col 8 (Main) and col 9
(Total) and they're equal; we use col 9 with a col-8 fallback.

  1  DATE
  2  SUPPLIER (filled on purchase rows: MICMAKIN, AKURE, BAKERY, ...)
  3  OPENING STOCK MAIN
  4  TOTAL OPENING STOCK (LITRES)
  5  PURCHASES / TRANSFER IN   -> diesel_added
  6  DIESEL TRANSFER OUT       (none present in this file)
  7  DIESEL AVAILABLE
  8  CLOSING STOCK - MAIN
  9  TOTAL CLOSING STOCK       -> diesel_level_actual
  10 CONSUMPTION IN LITRES
  12 DIESEL COST (NAIRA)
  13 GENERATOR RUNNING HOUR Opening
  14 GENERATOR RUNNING HOUR Closing
  15 HOUR RUN (HOURS)
  16 GENERATOR CONSUMPTION (LITRES/HOUR)   (sheet value; we recompute)
  17 NEPA METER Opening
  18 NEPA Closing

Purchases/distributions are NOT touched here (supply ledger is separate).

Usage:
  py scripts\import_ikare_cr.py                       # dry run
  py scripts\import_ikare_cr.py --apply               # write readings
  py scripts\import_ikare_cr.py --apply --recalc-baseline
  py scripts\import_ikare_cr.py --apply --replace     # sheet wins on overlapping dates

Reads SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY from .env / SupabaseCreds.env / env.
"""
import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
import datetime
from pathlib import Path

import openpyxl

FILE = Path(r"C:\Users\MakindeAribo\Downloads\IKARE CR DIESEL TEMPLATE YEAR 2026.xlsx")
STORE_NAME = "Ikare CR"          # must match generators.loc / locations.name
GEN_LOC_MATCH = "ikare cr"        # case-insensitive loc match to find the generator

MONTHLY_TABS = {
    "JAN 2026": (2026, 1), "FEB 2026": (2026, 2), "MARCH 2026": (2026, 3),
    "APRIL 2026": (2026, 4), "MAY 2026": (2026, 5), "JUNE 2026": (2026, 6),
}

COL = {
    "date": 1,
    "supplier": 2,
    "opening_total": 4,
    "purchases_in": 5,      # PURCHASES / TRANSFER IN  -> diesel_added
    "transfer_out": 6,      # DIESEL TRANSFER OUT (none in this file)
    "closing_main": 8,
    "closing_total": 9,     # TOTAL CLOSING STOCK -> diesel_level_actual
    "consumption_l": 10,
    "gen_h_open": 13,
    "gen_h_close": 14,
    "hour_run": 15,
    "consumption_rate": 16,
    "nepa_open": 17,
    "nepa_close": 18,
}


def load_env_file():
    for name in (".env", "SupabaseCreds.env", "supabase.env"):
        p = Path(name)
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
        return


class Supabase:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def _req(self, method, path, params=None, body=None, extra_headers=None):
        url = f"{self.url}/rest/v1/{path}"
        if params:
            url += "?" + urllib.parse.urlencode(params, doseq=True)
        data = json.dumps(body).encode("utf-8") if body is not None else None
        headers = dict(self.headers)
        if extra_headers:
            headers.update(extra_headers)
        req = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read()
                return json.loads(raw) if raw else []
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {e.read().decode('utf-8','replace')}") from e

    def select(self, table, columns="*", filters=None):
        params = {"select": columns}
        if filters:
            params.update(filters)
        return self._req("GET", table, params=params)

    def insert(self, table, rows):
        return self._req("POST", table, body=rows, extra_headers={"Prefer": "return=representation"})

    def delete(self, table, filters):
        return self._req("DELETE", table, params=filters, extra_headers={"Prefer": "return=minimal"})


def n(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return None if (isinstance(v, float) and v != v) else float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("#"):
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


def parse_readings():
    wb = openpyxl.load_workbook(FILE, data_only=True)
    out = []
    for tab, (year, month) in MONTHLY_TABS.items():
        if tab not in wb.sheetnames:
            print(f"  ! missing tab {tab!r}")
            continue
        ws = wb[tab]
        cnt = 0
        for r in range(1, ws.max_row + 1):
            d = ws.cell(r, COL["date"]).value
            if not isinstance(d, datetime.datetime) or d.year != year or d.month != month:
                continue
            closing = n(ws.cell(r, COL["closing_total"]).value)
            if closing is None or closing <= 0:
                closing = n(ws.cell(r, COL["closing_main"]).value)
            gh_close = n(ws.cell(r, COL["gen_h_close"]).value)
            if not ((closing is not None and closing > 0) or (gh_close is not None and gh_close > 0)):
                continue
            gh_open = n(ws.cell(r, COL["gen_h_open"]).value)
            cons = n(ws.cell(r, COL["consumption_l"]).value)
            hours = (gh_close - gh_open) if (gh_open is not None and gh_close is not None) else None
            rate = round(cons / hours, 2) if (cons is not None and hours and hours > 0) else None
            sup = ws.cell(r, COL["supplier"]).value
            sup = sup.strip() if isinstance(sup, str) and sup.strip() else None
            out.append({
                "date": d.date().isoformat(),
                "gen_h_open": gh_open,
                "gen_h_close": gh_close,
                "closing": closing,
                "purchases_in": n(ws.cell(r, COL["purchases_in"]).value),
                "consumption_l": cons,
                "consumption_rate": rate,
                "nepa_open": n(ws.cell(r, COL["nepa_open"]).value),
                "nepa_close": n(ws.cell(r, COL["nepa_close"]).value),
                "supplier": sup,
            })
            cnt += 1
        print(f"  {tab}: {cnt} rows")
    return out


def recalc_baseline(sb, gen_id):
    rows = sb._req("GET", "diesel_readings", params={
        "select": "date,gen_hours_opening,gen_hours_closing,diesel_level_actual,diesel_added",
        "generator_id": f"eq.{gen_id}", "order": "date.asc",
    })
    rates = []
    prev = None
    for r in rows:
        ho, hc, lvl = r.get("gen_hours_opening"), r.get("gen_hours_closing"), r.get("diesel_level_actual")
        added = r.get("diesel_added") or 0
        hrs = (hc - ho) if (ho is not None and hc is not None) else None
        if hrs and hrs > 0 and lvl is not None and prev is not None:
            actual = prev + added - lvl
            if actual > 0:
                rate = actual / hrs
                if 1 <= rate <= 100:
                    rates.append(rate)
        if lvl is not None:
            prev = lvl
    if not rates:
        print("  No valid rate pairs — skipping baseline.")
        return
    avg = sum(rates) / len(rates)
    flagged = sum(1 for r in rates if abs(r - avg) / avg > 0.20) / len(rates) * 100
    print(f"  Pairs: {len(rates)}  Avg: {avg:.2f} L/hr (min {min(rates):.2f}, max {max(rates):.2f})  ~{flagged:.1f}% >20% off mean")
    sb._req("POST", "generator_baselines", params={"on_conflict": "generator_id"},
            body=[{"generator_id": gen_id, "avg_litres_per_hour": round(avg, 2),
                   "baseline_readings_count": len(rates),
                   "last_calculated": datetime.datetime.now(datetime.timezone.utc).isoformat(),
                   "min_rate": round(min(rates), 2), "max_rate": round(max(rates), 2)}],
            extra_headers={"Prefer": "resolution=merge-duplicates,return=representation"})
    print("  Baseline upserted.")


def main(apply_mode, recalc_mode, replace_mode):
    load_env_file()
    url, key = os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY required.")
        sys.exit(1)
    sb = Supabase(url, key)

    print("=== PARSING IKARE CR ===")
    readings = parse_readings()
    print(f"  Total: {len(readings)} reading rows\n")

    print("=== SUPABASE LOOKUP ===")
    gens = sb.select("generators", columns="id,name,loc,asset_type")
    gen = next((g for g in gens if (g.get("loc") or "").strip().lower() == GEN_LOC_MATCH
                and (g.get("asset_type") or "generator") != "oven"), None)
    if not gen:
        gen = next((g for g in gens if GEN_LOC_MATCH in (g.get("loc") or "").strip().lower()
                    and (g.get("asset_type") or "generator") != "oven"), None)
    if not gen:
        print(f"ERROR: no generator with loc ~ {STORE_NAME!r}. Found locs:")
        for g in sorted(gens, key=lambda x: (x.get("loc") or "")):
            print(f"  - {g['name']!r} loc={g.get('loc')!r} asset_type={g.get('asset_type')!r}")
        sys.exit(1)
    gen_id = gen["id"]
    print(f"  Ikare CR generator: {gen['name']} (id={gen_id}, loc={gen.get('loc')})")
    existing = sb.select("diesel_readings", columns="id,date,submitted_by", filters={"generator_id": f"eq.{gen_id}"})
    existing_dates = {r["date"] for r in existing}
    print(f"  Existing readings for this generator: {len(existing_dates)}")

    sheet_dates = {r["date"] for r in readings}
    overlap = [r for r in existing if r["date"] in sheet_dates]

    if replace_mode:
        new_readings = readings  # the sheet is authoritative for every date it covers
        print(f"\n=== REPLACE MODE ===")
        print(f"  The sheet wins on overlapping dates. Will DELETE {len(overlap)} existing reading(s) "
              f"on sheet dates, then insert all {len(readings)}.")
        for r in sorted(overlap, key=lambda x: x["date"]):
            print(f"    - delete {r['date']} (app entry by {r.get('submitted_by')})")
    else:
        new_readings = [r for r in readings if r["date"] not in existing_dates]

    print(f"\n=== INSERT PLAN ===")
    print(f"  diesel_readings to insert: {len(new_readings)} "
          f"({'replace mode — all sheet rows' if replace_mode else f'skip {len(readings) - len(new_readings)} dupes'})")
    if new_readings:
        print(f"  Date range: {new_readings[0]['date']} .. {new_readings[-1]['date']}")
        tot_pur = sum(r["purchases_in"] or 0 for r in new_readings)
        print(f"  Diesel added (purchases/transfer-in) total: {tot_pur:,.0f} L")

    if not apply_mode:
        print("\nDRY RUN — no writes. Re-run with --apply to write.")
        if recalc_mode:
            print("\n=== BASELINE RECALC (Ikare CR) ===")
            recalc_baseline(sb, gen_id)
        return

    print("\n=== WRITING ===")
    if replace_mode and overlap:
        for r in overlap:
            sb.delete("diesel_readings", {"id": f"eq.{r['id']}"})
        print(f"  - Deleted {len(overlap)} overlapping reading(s).")
    if new_readings:
        payload = [{
            "generator_id": gen_id, "store_location": STORE_NAME, "date": r["date"],
            "gen_hours_opening": r["gen_h_open"], "gen_hours_closing": r["gen_h_close"],
            "diesel_level_actual": r["closing"], "diesel_added": r["purchases_in"] or 0,
            "consumption_litres": int(r["consumption_l"]) if r["consumption_l"] is not None else None,
            "consumption_rate": r["consumption_rate"],
            "nepa_meter_opening": r["nepa_open"], "nepa_meter_closing": r["nepa_close"],
            "submitted_by": None,
            "notes": f"Supplier: {r['supplier']}" if r["supplier"] else None,
        } for r in new_readings]
        BATCH = 100
        total = 0
        for i in range(0, len(payload), BATCH):
            total += len(sb.insert("diesel_readings", payload[i:i + BATCH]))
        print(f"  + Inserted {total} Ikare CR readings")
    print("\nDONE.")
    if recalc_mode:
        print("\n=== BASELINE RECALC (Ikare CR) ===")
        recalc_baseline(sb, gen_id)


if __name__ == "__main__":
    main("--apply" in sys.argv, "--recalc-baseline" in sys.argv, "--replace" in sys.argv)
