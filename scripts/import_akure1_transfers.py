#!/usr/bin/env python
r"""
FleetPro: Import Akure 1's historical diesel transfers from the Excel file.

Akure 1's monthly tabs record transfer-out LITRES in column F and (sometimes)
the receiving vehicles in column G as "NAME (litres), NAME (litres)". The
original readings import only preserved column G text in notes — and G was
mostly empty — so the transfer numbers never made it into the database.
This script reads the FILE directly:

  - If column G has "NAME (NN)" pairs -> one VEHICLE transfer per pair
    (matched to the vehicles table by plate/name when possible)
  - Else if column F has litres        -> one transfer, dest 'other'
    (destination unknown; label 'Unspecified (from sheet)')

Dedup: each row gets notes='sheet-akure1:<date>:<n>' — re-runs skip dates
already imported. Safe to run repeatedly.

Usage:
  py scripts\import_akure1_transfers.py            # dry run
  py scripts\import_akure1_transfers.py --apply     # write

Then run: recalc_baselines.py --apply  and  backfill_discrepancy_flags.py --apply
"""
import json
import os
import re
import sys
import datetime
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl

FILE = Path(r"C:\Users\MakindeAribo\Downloads\AKURE 1 DIESEL REPORT.xlsx")
STORE = "Akure 1"

MONTHLY_TABS = {
    "MARCH 2025": (2025, 3), "APRIL 2025": (2025, 4), "MAY 2025": (2025, 5),
    "JUNE 2025": (2025, 6), "JULY 2025": (2025, 7), "AUGUST 2025": (2025, 8),
    "SEPTEMBER 2025": (2025, 9), "OCTOBER 2025": (2025, 10), "NOVEMBER 2025": (2025, 11),
    "DECEMBER 2025": (2025, 12),
    "JANUARY 2026": (2026, 1), "FEBRUARY 2026": (2026, 2), "MARCH 2026": (2026, 3),
    "APRIL": (2026, 4), "MAY": (2026, 5),
}
COL_TRANSFER_OUT = 6   # DIESEL TRANSFER OUT (litres)
COL_RECEIVED_BY = 7    # DIESEL RECEIVED BY (text, often empty)


def load_env_file():
    for name in (".env", "SupabaseCreds.env", "supabase.env"):
        p = Path(name)
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
        return


class Supabase:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def _req(self, method, path, params=None, body=None, extra_headers=None):
        full = f"{self.url}/rest/v1/{path}"
        if params:
            full += "?" + urllib.parse.urlencode(params, doseq=True)
        data = json.dumps(body).encode("utf-8") if body is not None else None
        headers = dict(self.headers)
        if extra_headers:
            headers.update(extra_headers)
        req = urllib.request.Request(full, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read()
                return json.loads(raw) if raw else []
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {e.read().decode('utf-8','replace')}") from e

    def select_all(self, table, columns, filters=None, page=1000):
        out = []
        offset = 0
        while True:
            params = {"select": columns, "limit": str(page), "offset": str(offset)}
            if filters:
                params.update(filters)
            chunk = self._req("GET", table, params=params)
            out.extend(chunk)
            if len(chunk) < page:
                return out
            offset += page

    def insert(self, table, rows):
        return self._req("POST", table, body=rows, extra_headers={"Prefer": "return=representation"})


def n(v):
    if isinstance(v, (int, float)):
        return None if (isinstance(v, float) and v != v) else float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(",", "").strip())
        except ValueError:
            return None
    return None


def norm(s):
    return re.sub(r"[^A-Z0-9]", "", (s or "").upper())


def main(apply_mode):
    load_env_file()
    url, key = os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY required.")
        sys.exit(1)
    sb = Supabase(url, key)

    gens = sb.select_all("generators", "id,name,loc,asset_type")
    src = next((g for g in gens if (g.get("loc") or "").strip().lower() == STORE.lower() and g.get("asset_type") != "oven"), None)
    if not src:
        print(f"ERROR: no generator at loc '{STORE}'.")
        sys.exit(1)
    print(f"Source generator: {src['name']} (id={src['id']})")
    vehicles = sb.select_all("vehicles", "id,name,plate")

    existing = sb.select_all("diesel_transfers", "notes", filters={"notes": "like.sheet-akure1:*"})
    done = {e["notes"] for e in existing if e.get("notes")}
    print(f"Already imported from sheet: {len(done)} rows\n")

    wb = openpyxl.load_workbook(FILE, data_only=True)
    pair_re = re.compile(r"([^,()]+?)\s*\(\s*(\d+(?:\.\d+)?)\s*\)")
    new_rows = []
    unmatched = {}
    months = {}
    for tab, (yy, mm) in MONTHLY_TABS.items():
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        for r in range(1, ws.max_row + 1):
            d = ws.cell(r, 1).value
            if not isinstance(d, datetime.datetime) or d.year != yy or d.month != mm:
                continue
            litres = n(ws.cell(r, COL_TRANSFER_OUT).value)
            text = ws.cell(r, COL_RECEIVED_BY).value
            if not litres or litres <= 0:
                continue
            date = d.date().isoformat()
            items = []
            pairs = pair_re.findall(str(text)) if isinstance(text, str) else []
            if pairs:
                for label, l in pairs:
                    label = label.strip(" ,;|")
                    nl = norm(label)
                    veh = next((v for v in vehicles if nl and (nl in norm(v.get("plate")) or nl in norm(v.get("name")) or (norm(v.get("plate")) and norm(v.get("plate")) in nl))), None)
                    if not veh:
                        unmatched[label] = unmatched.get(label, 0) + 1
                    items.append({"dest_type": "vehicle", "dest_id": veh["id"] if veh else None,
                                  "dest_label": veh["name"] if veh else label, "litres": float(l)})
            else:
                items.append({"dest_type": "other", "dest_id": None,
                              "dest_label": "Unspecified (from sheet)", "litres": litres})
            for i, it in enumerate(items):
                tag = f"sheet-akure1:{date}:{i}"
                if tag in done:
                    continue
                new_rows.append({"date": date, "store_location": STORE, "source_generator_id": src["id"],
                                 "dest_type": it["dest_type"], "dest_id": it["dest_id"],
                                 "dest_label": it["dest_label"], "litres": it["litres"],
                                 "recorded_by": None, "notes": tag})
                mk = date[:7]
                m = months.setdefault(mk, {"n": 0, "litres": 0.0})
                m["n"] += 1
                m["litres"] += it["litres"]

    print("=== PLAN (Akure 1 transfers from sheet) ===")
    for mk, m in sorted(months.items()):
        print(f"  {mk}: {m['n']:>3} transfers  {m['litres']:>8.0f} L")
    print(f"  TOTAL: {len(new_rows)} transfers, {sum(t['litres'] for t in new_rows):.0f} L")
    if unmatched:
        print("\n  Vehicle labels not matched (kept as label):")
        for lbl, cnt in sorted(unmatched.items(), key=lambda x: -x[1]):
            print(f"    {lbl!r} x{cnt}")

    if not apply_mode:
        print("\nDRY RUN — no writes. Re-run with --apply to write.")
        return
    BATCH = 100
    total = 0
    for i in range(0, len(new_rows), BATCH):
        total += len(sb.insert("diesel_transfers", new_rows[i:i + BATCH]))
    print(f"\n+ Inserted {total} transfers.")
    print("Now run: recalc_baselines.py --apply  then  backfill_discrepancy_flags.py --apply")


if __name__ == "__main__":
    main("--apply" in sys.argv)
