#!/usr/bin/env python
r"""
FleetPro: Re-import AKURE 1 daily diesel readings - Mar 2025 to Jul 2026.

Source: "AKURE 1 DIESEL REPORT (3).xlsx" - monthly tabs MARCH 2025..JULY2026
(17 tabs; "APRIL"/"MAY" have no year = 2026, "JULY2026" has no space; MARCH
2025 has ~255 junk columns, harmless). Replaces the original Cowork-era
import (456 rows through May 30 2026) and extends through early July.

WIDE column layout (Akure 1's template has RECEIVED BY at col 7, shifting
everything right of TRANSFER OUT by one vs the narrow template):
  1 DATE | 2 SUPPLIER | 4 TOTAL OPENING | 5 PURCHASES/IN -> diesel_added
  6 TRANSFER OUT -> diesel_transfers (bare litres, no vehicle names in this
    version; dedup by (date,litres) against the 316 already-reconstructed
    rows) | 7 RECEIVED BY (empty) | 9/10 CLOSING (use 10, fallback 9)
  11 CONSUMPTION | 13/14 GEN HOURS | 15 HOUR RUN | 16 rate (recomputed)
  17/18 NEPA

Verified: opening - closing + purchases - transfer = consumption. NOTE:
Akure 1's HOUR-METER discipline is poor (e.g. 610 L in "0.2 h" -> 3,050
L/hr) - rates/baseline stay unreliable here by construction; the recalc's
sanity fallback keeps its baseline at the positive-rate mean.

Usage:
  py scripts\import_akure1.py                       # dry run
  py scripts\import_akure1.py --apply --replace     # sheet is authoritative

Then: recalc_baselines.py --apply -> backfill_discrepancy_flags.py --apply ->
sync_generator_hours.py --apply
"""
import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
import datetime
from pathlib import Path

import openpyxl

FILE = Path(r"C:\Users\MakindeAribo\Downloads\AKURE 1 DIESEL REPORT (3).xlsx")
STORE_NAME = "Akure 1"
GEN_LOC_MATCH = "akure 1"
DUPLICATE_APP_DATES = set()   # none needed for Akure 1

# tab -> (year, month); spans two calendar years
MONTH_TABS = {
    "MARCH 2025": (2025, 3), "APRIL 2025": (2025, 4), "MAY 2025": (2025, 5),
    "JUNE 2025": (2025, 6), "JULY 2025": (2025, 7), "AUGUST 2025": (2025, 8),
    "SEPTEMBER 2025": (2025, 9), "OCTOBER 2025": (2025, 10), "NOVEMBER 2025": (2025, 11),
    "DECEMBER 2025": (2025, 12), "JANUARY 2026": (2026, 1), "FEBRUARY 2026": (2026, 2),
    "MARCH 2026": (2026, 3), "APRIL": (2026, 4), "MAY": (2026, 5),
    "JUNE 2026": (2026, 6), "JULY2026": (2026, 7),
}

COL = {
    "supplier": 2, "opening_total": 4, "purchases_in": 5, "transfer_out": 6,
    "closing_main": 9, "closing_total": 10, "consumption_l": 11,
    "gen_h_open": 13, "gen_h_close": 14, "hour_run": 15, "consumption_rate": 16,
    "nepa_open": 17, "nepa_close": 18,
}


def load_env_file():
    for name in (".env", "SupabaseCreds.env", "supabase.env"):
        p = Path(name)
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
        return


class Supabase:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def _req(self, method, path, params=None, body=None, extra_headers=None):
        url = f"{self.url}/rest/v1/{path}"
        if params:
            url += "?" + urllib.parse.urlencode(params, doseq=True)
        data = json.dumps(body).encode("utf-8") if body is not None else None
        headers = dict(self.headers)
        if extra_headers:
            headers.update(extra_headers)
        req = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read()
                return json.loads(raw) if raw else []
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {e.read().decode('utf-8','replace')}") from e

    def select(self, table, columns="*", filters=None):
        params = {"select": columns}
        if filters:
            params.update(filters)
        return self._req("GET", table, params=params)

    def insert(self, table, rows):
        return self._req("POST", table, body=rows, extra_headers={"Prefer": "return=representation"})

    def delete(self, table, filters):
        return self._req("DELETE", table, params=filters, extra_headers={"Prefer": "return=minimal"})


def n(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return None if (isinstance(v, float) and v != v) else float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("#"):
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


def parse_readings():
    wb = openpyxl.load_workbook(FILE, data_only=True)
    out = []
    for tab, (year, month) in MONTH_TABS.items():
        if tab not in wb.sheetnames:
            print(f"  ! missing tab {tab!r}")
            continue
        ws = wb[tab]
        cnt = 0
        for r in range(1, ws.max_row + 1):
            d = ws.cell(r, 1).value
            if not isinstance(d, datetime.datetime) or d.year != year or d.month != month:
                continue
            closing = n(ws.cell(r, COL["closing_total"]).value)
            if closing is None or closing <= 0:
                closing = n(ws.cell(r, COL["closing_main"]).value)
            gh_close = n(ws.cell(r, COL["gen_h_close"]).value)
            # Real day = closing level OR closing gen-hours; excludes the July
            # trailing template artifact (blank closing -> cons = opening).
            if not ((closing is not None and closing > 0) or (gh_close is not None and gh_close > 0)):
                continue
            gh_open = n(ws.cell(r, COL["gen_h_open"]).value)
            cons = n(ws.cell(r, COL["consumption_l"]).value)
            hours = (gh_close - gh_open) if (gh_open is not None and gh_close is not None) else None
            rate = round(cons / hours, 2) if (cons is not None and hours and hours > 0.05) else None
            tx = n(ws.cell(r, COL["transfer_out"]).value) or 0
            sup = ws.cell(r, COL["supplier"]).value
            sup = sup.strip() if isinstance(sup, str) and sup.strip() else None
            notes = []
            if sup:
                notes.append(f"Supplier: {sup}")
            if tx > 0:
                notes.append(f"Transfer-out: {tx:.0f}")
            out.append({
                "date": d.date().isoformat(),
                "gen_h_open": gh_open, "gen_h_close": gh_close,
                "closing": round(closing, 1) if closing is not None else None,
                "purchases_in": n(ws.cell(r, COL["purchases_in"]).value),
                "consumption_l": round(cons, 1) if cons is not None else None,
                "consumption_rate": rate,
                "nepa_open": n(ws.cell(r, COL["nepa_open"]).value),
                "nepa_close": n(ws.cell(r, COL["nepa_close"]).value),
                "transfer_out": tx,
                "notes": "; ".join(notes) if notes else None,
            })
            cnt += 1
        print(f"  {tab:<14}: {cnt} rows")
    out.sort(key=lambda r: r["date"])
    return out


def main(apply_mode, replace_mode):
    load_env_file()
    url, key = os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY required.")
        sys.exit(1)
    sb = Supabase(url, key)

    print("=== PARSING AKURE 1 ===")
    readings = parse_readings()
    tot_tx = sum(r["transfer_out"] for r in readings)
    print(f"  Total: {len(readings)} rows | purchases {sum(r['purchases_in'] or 0 for r in readings):,.0f} L | transfers-out {tot_tx:,.0f} L\n")

    print("=== SUPABASE LOOKUP ===")
    gens = sb.select("generators", columns="id,name,loc,asset_type")
    gen = next((g for g in gens if (g.get("loc") or "").strip().lower() == GEN_LOC_MATCH
                and (g.get("asset_type") or "generator") != "oven"), None)
    if not gen:
        print(f"ERROR: no generator with loc == {STORE_NAME!r}.")
        sys.exit(1)
    gen_id = gen["id"]
    print(f"  Akure 1 generator: {gen['name']} (id={gen_id})")
    existing = sb.select("diesel_readings", columns="id,date,submitted_by,gen_hours_opening,gen_hours_closing",
                         filters={"generator_id": f"eq.{gen_id}"})
    print(f"  Existing readings: {len(existing)}")

    sheet_dates = {r["date"] for r in readings}
    overlap = [r for r in existing if r["date"] in sheet_dates or r["date"] in DUPLICATE_APP_DATES]

    if replace_mode:
        new_readings = readings
        print(f"\n=== REPLACE MODE ===")
        print(f"  Sheet wins on its dates + duplicate app dates {sorted(DUPLICATE_APP_DATES)}.")
        print(f"  Will DELETE {len(overlap)} existing reading(s), insert {len(readings)}.")
        app_over = [r for r in overlap if r.get("submitted_by")]
        if app_over:
            print(f"  ({len(app_over)} app-entered: {', '.join(sorted(r['date'] for r in app_over))})")
    else:
        existing_dates = {r["date"] for r in existing}
        new_readings = [r for r in readings if r["date"] not in existing_dates]

    print(f"\n=== PLAN ===")
    print(f"  Insert: {len(new_readings)} readings ({new_readings[0]['date']} .. {new_readings[-1]['date']})")
    tx_rows = [r for r in readings if r["transfer_out"] > 0]
    print(f"  diesel_transfers to upsert: {len(tx_rows)} ({tot_tx:.0f} L)")

    if not apply_mode:
        print("\nDRY RUN — no writes. Re-run with --apply --replace.")
        return

    print("\n=== WRITING ===")
    if replace_mode and overlap:
        for r in overlap:
            sb.delete("diesel_readings", {"id": f"eq.{r['id']}"})
        print(f"  - Deleted {len(overlap)} reading(s).")
    if new_readings:
        payload = [{
            "generator_id": gen_id, "store_location": STORE_NAME, "date": r["date"],
            "gen_hours_opening": r["gen_h_open"], "gen_hours_closing": r["gen_h_close"],
            "diesel_level_actual": r["closing"], "diesel_added": r["purchases_in"] or 0,
            "consumption_litres": r["consumption_l"],
            "consumption_rate": r["consumption_rate"],
            "nepa_meter_opening": r["nepa_open"], "nepa_meter_closing": r["nepa_close"],
            "submitted_by": None, "notes": r["notes"],
        } for r in new_readings]
        BATCH = 100
        total = 0
        for i in range(0, len(payload), BATCH):
            total += len(sb.insert("diesel_readings", payload[i:i + BATCH]))
        print(f"  + Inserted {total} Akure 1 readings")
    # Transfers: dedup by (date, store, litres) so re-runs are safe.
    existing_tx = sb.select("diesel_transfers", columns="date,store_location,litres",
                            filters={"store_location": f"eq.{STORE_NAME}"})
    have = {(t["date"], float(t["litres"])) for t in existing_tx}
    added_tx = 0
    for r in tx_rows:
        if (r["date"], float(r["transfer_out"])) in have:
            continue
        sb.insert("diesel_transfers", [{
            "date": r["date"], "store_location": STORE_NAME,
            "source_generator_id": gen_id, "dest_type": "other", "dest_id": None,
            "dest_label": "Transfer out (sheet)", "litres": r["transfer_out"],
            "notes": "From AKURE 1 DIESEL REPORT import", "recorded_by": None,
        }])
        added_tx += 1
    print(f"  + Inserted {added_tx} diesel_transfers rows")
    print("\nDONE. Next: recalc_baselines.py --apply -> backfill_discrepancy_flags.py --apply -> sync_generator_hours.py --apply")


if __name__ == "__main__":
    main("--apply" in sys.argv, "--replace" in sys.argv)
