#!/usr/bin/env python
r"""
FleetPro: Import diesel readings for the 7 bakery stores (ovens + generators).

Assets per store (must exist before running — see setup_bakery_assets.py):
  Oven only:        Akure Bakery, Ado Bakery, Ikare Bakery, Okitipupa Bakery
  Oven + generator: Ondo Bakery, Owo Bakery, Oye Bakery

Template layouts (columns relative to the DATE column, auto-offset per tab):
  OVEN:      supplier=2, opening=3, in=5, out=6, closing=8, consumption=10,
             BATCHES=12, rate(L/batch)=13. No hour meter, no NEPA.
  GENERATOR: same as the OWO CR template — in=5, out=6, closing=8,
             consumption=10, genOpen=12, genClose=13, hourRun=14, rate=15,
             nepaOpen=16, nepaClose=17.

Quirks handled automatically:
  - Some tabs are shifted right (e.g. IKARE 'JUNE 2026' starts in column B):
    the DATE column is auto-detected and all columns offset from it.
  - OYE BAKERY's generator hour meter counts MINUTES: detected per tab by
    comparing (close-open) to the HOUR RUN column; meters are divided by 60.

Usage:
  py scripts\import_bakeries.py                       # dry run
  py scripts\import_bakeries.py --apply               # write readings
  py scripts\import_bakeries.py --apply --recalc-baseline   # + gen baselines

Reads SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY from .env / SupabaseCreds.env / env.
"""
import json
import os
import sys
import datetime
import statistics
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl

DL = r"C:\Users\MakindeAribo\Downloads"

FILES = [
    {"file": f"{DL}\\Daily Diesel Tracker Bakery Oven AKURE BAKERY 2026.xlsx", "store": "Akure Bakery", "asset": "oven",
     "tabs": {"JANUARY 2026": (2026, 1), "FEBRUARY 2026": (2026, 2), "MARCH 2026": (2026, 3), "APRIL2026": (2026, 4), "MAY2026": (2026, 5), "JUNE2026": (2026, 6)}},
    {"file": f"{DL}\\ADO BAKERY DAILY DIESEL TEMPLATE 2026..xlsx", "store": "Ado Bakery", "asset": "oven",
     "tabs": {"JULY,2025": (2025, 7), "AUGUST 2025": (2025, 8), "SEPTEMBER,25": (2025, 9), "OCTOBER 25": (2025, 10), "NOVEMBER 2025": (2025, 11),
              "JANUARY 2026": (2026, 1), "FEB 2026": (2026, 2), "MARCH 2026": (2026, 3), "APRIL 2026": (2026, 4), "MAY 2026": (2026, 5), "JUNE,2026": (2026, 6)}},
    {"file": f"{DL}\\IKARE BAKERY DIESEL TEMPLATE FOR 2026.xlsx", "store": "Ikare Bakery", "asset": "oven",
     "tabs": {"JAN,2026": (2026, 1), "FEB 2026": (2026, 2), "MARCH 2026": (2026, 3), "APRIL 2026": (2026, 4), "MAY 2026": (2026, 5), "JUNE 2026": (2026, 6)}},
    {"file": f"{DL}\\Okpa bakery Daily Diesel Tracker Bakery Oven.xlsx", "store": "Okitipupa Bakery", "asset": "oven",
     "tabs": {"JULY 2025": (2025, 7), "AUGUST 2025": (2025, 8), "SEPT 2025": (2025, 9), "OCT 2025": (2025, 10), "NOV 2025": (2025, 11), "DEC 2025": (2025, 12),
              "JAN 2026": (2026, 1), "FEB 2026": (2026, 2), "MARCH.2026": (2026, 3), "APRIL.2026": (2026, 4), "MAY,26": (2026, 5), "JUNE": (2026, 6)}},
    {"file": f"{DL}\\OYE BAKERY GENERATOR DIESEL TRACKER, (1).xlsx", "store": "Oye Bakery", "asset": "generator",
     "tabs": {"JANUARY 2026": (2026, 1), "FEBRUARY 2026": (2026, 2), "MARCH 2026": (2026, 3), "APRIL 2026": (2026, 4), "MAY 2026": (2026, 5), "JUNE 2026": (2026, 6)}},
    {"file": f"{DL}\\OYE BAKERY DAILY Diesel OVEN TrackerR,.xlsx", "store": "Oye Bakery", "asset": "oven",
     "tabs": {"JANUARY 2026": (2026, 1), "FEBRUARY 2026": (2026, 2), "MARCH 2026": (2026, 3), "APRIL 2026": (2026, 4), "MAY 2026": (2026, 5), "JUNE 2026": (2026, 6)}},
    {"file": f"{DL}\\Daily Diesel Tracker OWO Bakery Oven-2026.xlsx", "store": "Owo Bakery", "asset": "oven",
     "tabs": {"JAN 2026": (2026, 1), "FEBRUARY 2026": (2026, 2), "MARCH 2026": (2026, 3), "APRIL 2026": (2026, 4), "MAY 2026": (2026, 5), "JUNE2026": (2026, 6)}},
    {"file": f"{DL}\\Daily Diesel Tracker for OWO BAKERY GENERATOR 2026.xlsx", "store": "Owo Bakery", "asset": "generator",
     "tabs": {"JAN 2026": (2026, 1), "FEBRUARY 2026": (2026, 2), "MARCH 2026": (2026, 3), "APRIL": (2026, 4), "MAY": (2026, 5), "JUNE": (2026, 6)}},
    {"file": f"{DL}\\ONDO BAKERY OVEN DIESEL 2026.xlsx", "store": "Ondo Bakery", "asset": "oven",
     "tabs": {"JANUARY  2026": (2026, 1), "feb 2026": (2026, 2), "march 2026": (2026, 3), "APRIL 2026": (2026, 4), "MAY 2026": (2026, 5), "JUNE 2026": (2026, 6)}},
    {"file": f"{DL}\\Generator Daily Diesel Tracker ondo bakery  (1).xlsx", "store": "Ondo Bakery", "asset": "generator",
     "tabs": {"JAN 2026 DAILY DIESEL TRACKER": (2026, 1), "feb 2026": (2026, 2), "march 2026": (2026, 3), "APRIL 2026": (2026, 4), "MAY 2026": (2026, 5), "JUNE 2026": (2026, 6)}},
]

# Column maps RELATIVE to the detected date column (date = 1)
OVEN_COL = {"supplier": 2, "purchases_in": 5, "transfer_out": 6, "closing_main": 8, "consumption_l": 10, "batches": 12, "rate": 13}
GEN_COL = {"supplier": 2, "purchases_in": 5, "transfer_out": 6, "closing_main": 8, "consumption_l": 10,
           "gen_open": 12, "gen_close": 13, "hour_run": 14, "rate": 15, "nepa_open": 16, "nepa_close": 17}


def load_env_file():
    for name in (".env", "SupabaseCreds.env", "supabase.env"):
        p = Path(name)
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
        return


class Supabase:
    def __init__(self, url, key):
        self.url = url.rstrip("/")
        self.headers = {"apikey": key, "Authorization": f"Bearer {key}", "Content-Type": "application/json"}

    def _req(self, method, path, params=None, body=None, extra_headers=None):
        full = f"{self.url}/rest/v1/{path}"
        if params:
            full += "?" + urllib.parse.urlencode(params, doseq=True)
        data = json.dumps(body).encode("utf-8") if body is not None else None
        headers = dict(self.headers)
        if extra_headers:
            headers.update(extra_headers)
        req = urllib.request.Request(full, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req) as resp:
                raw = resp.read()
                return json.loads(raw) if raw else []
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {e.read().decode('utf-8','replace')}") from e

    def select_all(self, table, columns, filters=None, order=None, page=1000):
        out = []
        offset = 0
        while True:
            params = {"select": columns, "limit": str(page), "offset": str(offset)}
            if filters:
                params.update(filters)
            if order:
                params["order"] = order
            chunk = self._req("GET", table, params=params)
            out.extend(chunk)
            if len(chunk) < page:
                return out
            offset += page

    def insert(self, table, rows):
        return self._req("POST", table, body=rows, extra_headers={"Prefer": "return=representation"})


def n(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return None if (isinstance(v, float) and v != v) else float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s.startswith("#"):
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


def find_date_col(ws):
    """Detect which column holds the dates (some tabs are shifted right)."""
    for c in range(1, 5):
        for r in range(1, 25):
            if isinstance(ws.cell(r, c).value, datetime.datetime):
                return c
    return None


def parse_tab(ws, year, month, asset):
    date_col = find_date_col(ws)
    if date_col is None:
        return [], "no date column found"
    off = date_col - 1
    colmap = OVEN_COL if asset == "oven" else GEN_COL
    raw = []
    for r in range(1, ws.max_row + 1):
        d = ws.cell(r, date_col).value
        if not isinstance(d, datetime.datetime) or d.year != year or d.month != month:
            continue
        rec = {"date": d.date().isoformat()}
        for key, rel in colmap.items():
            v = ws.cell(r, rel + off).value
            rec[key] = (v.strip() if isinstance(v, str) else v) if key == "supplier" else n(v)
        raw.append(rec)
    # Keep only real, completed days
    if asset == "oven":
        rows = [r for r in raw if (r["closing_main"] is not None and r["closing_main"] > 0) or (r["batches"] or 0) > 0 or (r["consumption_l"] or 0) > 0]
    else:
        rows = [r for r in raw if (r["closing_main"] is not None and r["closing_main"] > 0) or (r["gen_close"] or 0) > 0]
        # Minutes-meter detection: if (close-open)/hour_run clusters near 60, the meter counts minutes
        ratios = []
        for r in rows:
            if r["gen_open"] and r["gen_close"] and (r["hour_run"] or 0) > 0.2:
                delta = r["gen_close"] - r["gen_open"]
                if delta > 0:
                    ratios.append(delta / r["hour_run"])
        if ratios and 30 < statistics.median(ratios) < 90:
            for r in rows:
                if r["gen_open"] is not None:
                    r["gen_open"] = round(r["gen_open"] / 60, 2)
                if r["gen_close"] is not None:
                    r["gen_close"] = round(r["gen_close"] / 60, 2)
            return rows, "minutes meter -> /60"
    return rows, None


def main(apply_mode, recalc_mode):
    load_env_file()
    url, key = os.environ.get("SUPABASE_URL"), os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("ERROR: SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY required.")
        sys.exit(1)
    sb = Supabase(url, key)

    gens = sb.select_all("generators", "id,name,loc,asset_type")

    def find_asset(store, kind):
        rows = [g for g in gens if (g.get("loc") or "").strip().lower() == store.lower()]
        if kind == "oven":
            return next((g for g in rows if g.get("asset_type") == "oven"), None)
        return next((g for g in rows if g.get("asset_type") != "oven"), None)

    plans = []
    total_new = 0
    print("=== PARSING + LOOKUP ===")
    for cfg in FILES:
        fname = cfg["file"].split("\\")[-1]
        asset_row = find_asset(cfg["store"], cfg["asset"])
        if not asset_row:
            print(f"\n! {cfg['store']} ({cfg['asset']}): NO MATCHING ASSET in DB — run setup_bakery_assets.py --apply first. Skipping {fname}")
            continue
        if not Path(cfg["file"]).exists():
            print(f"\n! File not found: {cfg['file']} — skipping")
            continue
        wb = openpyxl.load_workbook(cfg["file"], data_only=True)
        all_rows = []
        print(f"\n{cfg['store']} [{cfg['asset']}] -> {asset_row['name']} (id={asset_row['id']})   file: {fname}")
        for tab, (yy, mm) in cfg["tabs"].items():
            if tab not in wb.sheetnames:
                print(f"  ! tab missing: {tab!r}")
                continue
            rows, note = parse_tab(wb[tab], yy, mm, cfg["asset"])
            print(f"  {tab}: {len(rows)} rows{('  [' + note + ']') if note else ''}")
            all_rows.extend(rows)
        existing = {r["date"] for r in sb.select_all("diesel_readings", "date", filters={"generator_id": f"eq.{asset_row['id']}"})}
        new_rows = [r for r in all_rows if r["date"] not in existing]
        total_new += len(new_rows)
        if all_rows:
            print(f"  => parsed {len(all_rows)}, new {len(new_rows)} (skip {len(all_rows)-len(new_rows)} dupes)  range {min(r['date'] for r in all_rows)} .. {max(r['date'] for r in all_rows)}")
        plans.append((cfg, asset_row, new_rows))

    print(f"\n=== INSERT PLAN ===  total new readings: {total_new}")
    if not apply_mode:
        print("\nDRY RUN — no writes. Re-run with --apply to write.")
        return

    print("\n=== WRITING ===")
    for cfg, asset_row, new_rows in plans:
        if not new_rows:
            continue
        payload = []
        for r in new_rows:
            notes = []
            if r.get("supplier"):
                notes.append(f"Supplier: {r['supplier']}")
            if r.get("transfer_out"):
                notes.append(f"Transfer-out: {r['transfer_out']}")
            base = {
                "generator_id": asset_row["id"],
                "store_location": cfg["store"],
                "date": r["date"],
                "diesel_level_actual": r["closing_main"],
                "diesel_added": r["purchases_in"] or 0,
                "consumption_litres": int(r["consumption_l"]) if r["consumption_l"] is not None else None,
                "consumption_rate": r["rate"],
                "submitted_by": None,
                "notes": " | ".join(notes) if notes else None,
            }
            if cfg["asset"] == "oven":
                base["batches_produced"] = r["batches"]
            else:
                base["gen_hours_opening"] = r["gen_open"]
                base["gen_hours_closing"] = r["gen_close"]
                base["nepa_meter_opening"] = r["nepa_open"]
                base["nepa_meter_closing"] = r["nepa_close"]
            payload.append(base)
        BATCH = 100
        total = 0
        for i in range(0, len(payload), BATCH):
            total += len(sb.insert("diesel_readings", payload[i:i + BATCH]))
        print(f"  + {cfg['store']} [{cfg['asset']}]: inserted {total}")
    print("\nDONE.")

    if recalc_mode:
        print("\n=== BASELINE RECALC (bakery generators only — ovens have no hour meter) ===")
        for cfg, asset_row, _ in plans:
            if cfg["asset"] != "generator":
                continue
            rows = sb.select_all("diesel_readings",
                                 "date,gen_hours_opening,gen_hours_closing,diesel_level_actual,diesel_added",
                                 filters={"generator_id": f"eq.{asset_row['id']}"}, order="date.asc")
            rates = []
            prev = None
            for r in rows:
                ho, hc, lvl = r.get("gen_hours_opening"), r.get("gen_hours_closing"), r.get("diesel_level_actual")
                added = r.get("diesel_added") or 0
                hrs = (hc - ho) if (ho is not None and hc is not None) else None
                if hrs and hrs > 0 and lvl is not None and prev is not None:
                    actual = prev + added - lvl
                    if actual > 0:
                        rate = actual / hrs
                        if 1 <= rate <= 100:
                            rates.append(rate)
                if lvl is not None:
                    prev = lvl
            if not rates:
                print(f"  [{asset_row['name']}] no valid pairs — skipped")
                continue
            avg = sum(rates) / len(rates)
            flagged = sum(1 for x in rates if abs(x - avg) / avg > 0.20) / len(rates) * 100
            sb._req("POST", "generator_baselines", params={"on_conflict": "generator_id"},
                    body=[{"generator_id": asset_row["id"], "avg_litres_per_hour": round(avg, 2),
                           "baseline_readings_count": len(rates),
                           "last_calculated": datetime.datetime.now(datetime.timezone.utc).isoformat(),
                           "min_rate": round(min(rates), 2), "max_rate": round(max(rates), 2)}],
                    extra_headers={"Prefer": "resolution=merge-duplicates,return=representation"})
            print(f"  [{asset_row['name']}] pairs={len(rates)} avg={avg:.2f} L/hr (min {min(rates):.2f}, max {max(rates):.2f}) ~{flagged:.1f}% >20% off — upserted")


if __name__ == "__main__":
    main("--apply" in sys.argv, "--recalc-baseline" in sys.argv)
